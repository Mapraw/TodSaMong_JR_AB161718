import streamlit as st
import pandas as pd
import requests
import openpyxl
from io import BytesIO
from datetime import datetime
import re

# Page Config
st.set_page_config(page_title="AB16-18 Proposal Data Investigation", layout="wide")

# Backend URL
BACKEND_URL = "http://localhost:8888"

# --- Connectivity Check ---
try:
    check = requests.get(f"{BACKEND_URL}/")
    is_online = (check.status_code == 200)
except:
    is_online = False

# --- Custom Styles ---
st.markdown("""
    <style>
    /* Sidebar Styling */
    section[data-testid="stSidebar"] { background-color: #f8f9fa; border-right: 1px solid #e9ecef; }
    div[data-testid="stSidebar"] div[role="radiogroup"] label {
        padding: 10px 15px !important; border-radius: 8px !important; margin-bottom: 4px !important;
        border: 1px solid transparent !important; transition: all 0.2s ease;
    }
    div[data-testid="stSidebar"] div[role="radiogroup"] label:hover { background-color: #e9ecef !important; }
    div[data-testid="stSidebar"] div[role="radiogroup"] label[data-baseweb="radio"] > div:first-child { display: none !important; }
    div[data-testid="stSidebar"] div[role="radiogroup"] label[aria-checked="true"] { background-color: #1a73e8 !important; border-color: #1a73e8 !important; }
    div[data-testid="stSidebar"] div[role="radiogroup"] label[aria-checked="true"] p { color: white !important; font-weight: 600 !important; }

    /* Tab Styling */
    .stTabs [data-baseweb="tab-list"] { gap: 8px; background-color: #f1f3f4; padding: 8px; border-radius: 12px; }
    .stTabs [data-baseweb="tab"] { background-color: white; border-radius: 8px; padding: 8px 16px; border: 1px solid #dee2e6; font-weight: 500; }
    .stTabs [aria-selected="true"] { background-color: #1a73e8 !important; color: white !important; }
    
    .status-badge { padding: 4px 12px; border-radius: 16px; font-weight: 600; font-size: 0.85rem; display: inline-block; }
    .status-valid { background-color: #e6f4ea; color: #1e8e3e; border: 1px solid #1e8e3e; }
    .status-invalid { background-color: #fce8e6; color: #d93025; border: 1px solid #d93025; }
    .status-incomplete { background-color: #fef7e0; color: #f9ab00; border: 1px solid #f9ab00; }
    
    .verdict-box { padding: 24px; border-radius: 12px; margin-top: 20px; border: 2px solid; }
    .verdict-pass { background-color: #e6f4ea; border-color: #1e8e3e; color: #1e8e3e; }
    .verdict-fail { background-color: #fce8e6; border-color: #d93025; color: #d93025; }
    .verdict-warn { background-color: #fef7e0; border-color: #f9ab00; color: #f9ab00; }
    </style>
    """, unsafe_allow_html=True)

# --- Session State ---
if "processed_data" not in st.session_state: st.session_state.processed_data = []
if "preview_data" not in st.session_state: st.session_state.preview_data = []

# --- Helper Functions ---
def get_category_label(cat):
    mapping = {
        "AB18": "AB18 (Low Voltage)",
        "AB17": "AB17 (XLPE)",
        "AB16_TERM": "AB16 (Terminations)",
        "AB16_CLEAT": "AB16 (Cleats)",
    }
    return mapping.get(cat, cat)

def render_download_button(data, label, filename, key, category="ALL", use_sidebar=False):
    if not data: return
    container = st.sidebar if use_sidebar else st
    try:
        resp = requests.post(f"{BACKEND_URL}/generate-report?category={category}", json=data)
        if resp.status_code == 200:
            container.download_button(f"📥 {label}", data=resp.content, file_name=f"{filename}_{datetime.now().strftime('%Y%b%d-%H%M')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key, use_container_width=True)
        else: 
            container.error(f"Report failed (Status {resp.status_code}): {resp.text}")
            print(f"[DEBUG] Report failed with {resp.status_code}: {resp.text}")
    except Exception as e: container.error(f"Error: {e}")

def render_overall_summary(results, category="ALL"):
    if not results: return
    unique_bidders = sorted(list(set(b.get("Bidder") for b in results if b.get("Bidder"))))
    if not unique_bidders:
        return
    all_equip, seen_eq = [], set()
    unwanted = ["TOTAL PRICE", "SUMMARY", "TRANSPORTATION", "GRAND TOTAL", "PRICE FOR SCHEDULE", "INSTALLATION", "PRICE OF SCHEDULE"]
    for b in results:
        sources = b.get("equip_order", []) + list(b.get("Equipment", {}).keys())
        for eq in sources:
            if eq not in seen_eq and not any(t in eq.upper() for t in unwanted): all_equip.append(eq); seen_eq.add(eq)
    
    # Sort all_equip: AB18 first, then AB17, then AB16
    def get_eq_sort_key(eq):
        item_id = eq.split()[0]
        match = re.search(r"(\d*)AB(16|17|18)(?:\D*(\d+))?", item_id, re.I)
        if match:
            prefix_num = int(match.group(1)) if match.group(1) else 0
            cat_num = int(match.group(2))
            sub_num = int(match.group(3)) if match.group(3) else 0
            cat_order = {18: 1, 17: 2, 16: 3}.get(cat_num, 99)
            return (cat_order, prefix_num, sub_num, eq)
        else:
            return (99, 0, 0, eq)
            
    all_equip = sorted(all_equip, key=get_eq_sort_key)
    
    summary_data = []
    for eq in all_equip:
        row = {"Item ID": eq.split()[0], "Description": " ".join(eq.split()[1:])}
        for b_name in unique_bidders:
            b_entries = [b for b in results if b.get("Bidder") == b_name]
            status = "N/A"; found = False; all_pass = True
            for entry in b_entries:
                if eq in entry.get("Equipment", {}):
                    res_list = entry["Equipment"][eq]
                    if res_list:
                        found = True
                        if not all(any(ok in str(r.get("result", "")).upper() for ok in ["PASS", "VALID", "ผ่าน"]) and "ไม่ผ่าน" not in str(r.get("result", "")) for r in res_list):
                            all_pass = False; break
            if found: row[b_name] = "Pass" if all_pass else "Fail"
            else: row[b_name] = "Not Offered"
        summary_data.append(row)

    if category == "AB18":
        verdict_row = {"Item ID": "", "Description": "OVERALL VERDICT"}
        for b_name in unique_bidders:
            b_entries = [b for b in results if b.get("Bidder") == b_name]
            if not b_entries: verdict_row[b_name] = "Not Offered"
            else:
                if any(b.get("Status") in ["INVALID", "INCOMPLETE"] for b in b_entries): verdict_row[b_name] = "Fail"
                else: verdict_row[b_name] = "Pass"
        summary_data.append(verdict_row)

    df_sum = pd.DataFrame(summary_data)
    def color_status(val):
        if val == "Pass": return 'background-color: #e6f4ea; color: #1e8e3e'
        if val == "Fail": return 'background-color: #fce8e6; color: #d93025'
        return ''
    st.markdown("#### 📊 Summary Overview")
    st.dataframe(df_sum.style.map(color_status, subset=df_sum.columns.intersection(unique_bidders)), use_container_width=True, hide_index=True)

def render_bidder_analysis(b_data):
    status = b_data.get('Status', 'UNKNOWN')
    category = str(b_data.get("Category", ""))
    is_ab16 = category.startswith("AB16")
    s_class = "status-valid" if status == 'VALID' else "status-invalid" if status == 'INVALID' else "status-incomplete"
    st.markdown(f"### 🏢 {b_data['Bidder']}")
    c1, c2 = st.columns([1, 3])
    with c1:
        st.markdown(f"**Overall Status:** <span class='status-badge {s_class}'>{status}</span>", unsafe_allow_html=True)
        st.write(f"**Schedule:** {b_data.get('Schedule', 'N/A')}"); st.write(f"**Items Found:** {b_data.get('ItemNoSummary', 'N/A')}")
        if b_data.get('Issues'):
            st.error("**Flags Identified:**")
            for iss in b_data['Issues']: st.markdown(f"- {iss}")
    if is_ab16:
        with c2:
            for eq, mfrs in b_data.get('Equipment', {}).items():
                if any(t in eq.upper() for t in ["TOTAL PRICE", "SUMMARY", "TRANSPORTATION", "GRAND TOTAL", "PRICE FOR SCHEDULE", "INSTALLATION", "PRICE OF SCHEDULE"]):
                    continue
                with st.expander(f"Item: {eq}", expanded=True):
                    if not mfrs:
                        st.warning("No data found.")
                    else:
                        df = pd.DataFrame(mfrs).rename(columns={"manufacturer": "Manufacturer", "country": "Country", "standard": "Standard", "type_model": "Type/Model", "od_range": "OD Range", "result": "Result", "comment": "Comment"})
                        cols = [c for c in df.columns if c in ["Manufacturer", "Country", "Standard", "Type/Model", "OD Range", "Result", "Comment"]]
                        st.dataframe(df[cols] if cols else df, hide_index=True, use_container_width=True)

                        for i, row in enumerate(mfrs):
                            cmt = str(row.get("comment", ""))
                            if cmt and cmt not in ["Pass", "N/A", ""]:
                                mfr_name = row.get("manufacturer", "Unknown") or "Unknown"
                                country = row.get("country", "")
                                type_model = row.get("type_model", "")

                                label_parts = [mfr_name]
                                if country:
                                    label_parts.append(f"({country})")
                                if type_model:
                                    label_parts.append(f"- Model: {type_model}")

                                st.markdown(f"**Manufacturer Option: {' ' .join(label_parts)}**")

                                parts = cmt.split("\n\n")
                                if len(parts) >= 2:
                                    st.error(parts[0])
                                    st.info("\n\n".join(parts[1:]))
                                else:
                                    if "Ã¢Å“â€¦" in cmt:
                                        st.success(cmt)
                                    elif row.get("result", "Fail") == "Fail":
                                        st.error(cmt)
                                    else:
                                        st.info(cmt)
                st.markdown("---")
        return
    with c2:
        for eq, mfrs in b_data.get('Equipment', {}).items():
            if any(t in eq.upper() for t in ["TOTAL PRICE", "SUMMARY", "TRANSPORTATION", "GRAND TOTAL", "PRICE FOR SCHEDULE", "INSTALLATION", "PRICE OF SCHEDULE"]): continue
            with st.expander(f"📦 {eq}", expanded=True):
                if not mfrs: st.warning("No data found.")
                else:
                    df = pd.DataFrame(mfrs).rename(columns={"manufacturer": "Manufacturer", "country": "Country", "standard": "Standard", "type_model": "Type/Model", "od_range": "OD Range", "result": "Result", "comment": "Comment"})
                    cols = [c for c in df.columns if c in ["Manufacturer", "Country", "Standard", "Type/Model", "OD Range", "Result", "Comment"]]
                    st.dataframe(df[cols] if cols else df, hide_index=True, use_container_width=True)
                    
                    # Render detailed Qwerty-style comments
                    for i, row in enumerate(mfrs):
                        cmt = str(row.get("comment", ""))
                        if cmt and cmt not in ["Pass", "N/A", ""]:
                            # Display label to identify which manufacturer/model this comment is for
                            mfr_name = row.get("manufacturer", "Unknown") or "Unknown"
                            country = row.get("country", "")
                            type_model = row.get("type_model", "")
                            
                            label_parts = [mfr_name]
                            if country:
                                label_parts.append(f"({country})")
                            if type_model:
                                label_parts.append(f"- Model: {type_model}")
                            
                            st.markdown(f"**Manufacturer Option: {' '.join(label_parts)}**")
                            
                            # Split status and suggestion parts if delimited by double newline
                            parts = cmt.split("\n\n")
                            if len(parts) >= 2:
                                st.error(parts[0])
                                st.info("\n\n".join(parts[1:]))
                            else:
                                if "✅" in cmt:
                                    st.success(cmt)
                                elif row.get("result", "Fail") == "Fail":
                                    st.error(cmt)
                                else:
                                    st.info(cmt)
    st.divider()
    v_class = "verdict-pass" if status == "VALID" else "verdict-fail" if status == "INVALID" else "verdict-warn"
    v_icon = "✅" if status == "VALID" else "❌" if status == "INVALID" else "⚠️"
    st.markdown(f'<div class="verdict-box {v_class}"><h3 style="margin-top:0">{v_icon} Overall Verdict for {b_data["Bidder"]}</h3><p>Status: <b>{status}</b></p></div>', unsafe_allow_html=True)

def build_missing_bidder_warning(entries):
    files = []
    for entry in entries:
        for iss in entry.get("Issues", []):
            match = re.search(r"Forgot to fill bidder in file (.+)", str(iss).strip())
            if match:
                filename = match.group(1).strip()
                if filename and filename not in files:
                    files.append(filename)
    if files:
        return f"Please fill bidder name in file: {', '.join(files)}"
    return "Please fill bidder name before running the investigation for this file."

def merge_bidder_entries(entries, merged_category=""):
    if not entries:
        return {}

    merged = {
        "Bidder": entries[0].get("Bidder", ""),
        "Category": merged_category or entries[0].get("Category", ""),
        "Status": "VALID",
        "Schedule": ", ".join(sorted(list({str(e.get("Schedule", "")).strip() for e in entries if str(e.get("Schedule", "")).strip()}))),
        "ItemNoSummary": ", ".join(sorted(list({part.strip() for e in entries for part in str(e.get("ItemNoSummary", "")).split(",") if part.strip()}))),
        "Issues": [],
        "Equipment": {},
    }

    issue_set = []
    for entry in entries:
        entry_status = entry.get("Status", "VALID")
        if entry_status == "INVALID":
            merged["Status"] = "INVALID"
        elif entry_status == "INCOMPLETE" and merged["Status"] == "VALID":
            merged["Status"] = "INCOMPLETE"

        for iss in entry.get("Issues", []):
            if iss not in issue_set:
                issue_set.append(iss)

        for eq, rows in entry.get("Equipment", {}).items():
            merged["Equipment"].setdefault(eq, [])
            for row in rows:
                if row not in merged["Equipment"][eq]:
                    merged["Equipment"][eq].append(row)

    merged["Issues"] = issue_set
    return merged

# --- Sidebar ---
st.sidebar.title("Navigation")
if is_online: st.sidebar.success("🟢 Backend Connected")
else: st.sidebar.error("🔴 Backend Offline")

page = st.sidebar.radio("Go to", ["Upload & Analyze", "AB18 (Low Voltage)", "AB17 (XLPE)", "AB16 (Terminations/Cleats)"])

# PERSISTENT CONSOLIDATED DOWNLOAD BUTTON
if st.session_state.processed_data:
    st.sidebar.divider()
    st.sidebar.markdown("### 📥 Global Reports")
    render_download_button(st.session_state.processed_data, "Download Consolidated", "AB16-AB18", "dl_side_con", category="ALL", use_sidebar=True)

# --- Common Bidder Order ---
all_processed_bidders = sorted(list(set(b.get("Bidder") for b in st.session_state.processed_data if b.get("Bidder"))))

# --- Pages ---
if page == "Upload & Analyze":
    st.title("📂 Proposal Data Investigation System")
    m1, m2, m3 = st.columns(3)
    m1.metric("Unique Bidders", len(all_processed_bidders)); m2.metric("Categories", "4"); m3.metric("System", "Online" if st.session_state.processed_data else "Ready")
    c1, c2 = st.columns(2)
    with c1: ps_files = st.file_uploader("Upload Price Schedules", type=["xlsx"], accept_multiple_files=True)
    with c2: p_files = st.file_uploader("Upload Proposal Data", type=["xlsx"], accept_multiple_files=True)
    
    if p_files:
        if st.button("🚀 Run Investigation", use_container_width=True, type="primary"):
            with st.spinner("Analyzing..."):
                payload = [("files", (f.name, f.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")) for f in p_files]
                if ps_files:
                    for f in ps_files: payload.append(("price_schedules", (f.name, f.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")))
                try:
                    resp = requests.post(f"{BACKEND_URL}/upload?category_filter=ALL", files=payload)
                    if resp.status_code == 200: st.session_state.processed_data = resp.json().get("data", []); st.session_state.preview_data = []; st.rerun()
                    else: st.error(f"Error: {resp.text}")
                except Exception as e: st.error(f"Connection Failed: {e}")
        
        # Preview Section (Minimizable)
        if p_files and not st.session_state.preview_data:
            prev = []
            for pf in p_files:
                try:
                    # Use pandas to read the file first to check if it's readable at all
                    # This is more robust than openpyxl for checking validity
                    file_bytes = pf.getvalue()
                    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                    
                    for sn in wb.sheetnames:
                        try:
                            ws = wb[sn]
                            # Detect Category
                            a1 = str(ws['A1'].value or "").upper()
                            a2 = str(ws['A2'].value or "").upper()
                            
                            cat = "Unknown"
                            if "LOW VOLTAGE" in a2: cat = "AB18"
                            elif "XLPE POWER" in a1: cat = "AB17"
                            elif "TERMINATIONS" in a1: cat = "AB16_TERM"
                            elif "CLEATS" in a1: cat = "AB16_CLEAT"
                            
                            # EXTRACT BIDDER correctly using the helper
                            bidder = "Unknown"
                            for cell in ["C6", "C11", "C5", "C7", "C8", "C9", "C10"]:
                                val = str(ws[cell].value or "").strip()
                                if val and "MANUFACTURER" not in val.upper() and "BIDDER" not in val.upper() and len(val) > 2:
                                    bidder = val
                                    break
                                    
                            prev.append({
                                "Bidder": bidder,
                                "File": pf.name,
                                "Sheet": sn,
                                "Detected Category": get_category_label(cat)
                            })
                        except Exception as e:
                            prev.append({"Bidder": f"Sheet Error: {str(e)[:50]}", "File": pf.name, "Sheet": sn, "Detected Category": "N/A"})
                except Exception as e:
                    prev.append({"Bidder": f"File Error: {str(e)[:50]}", "File": pf.name, "Sheet": "All Sheets", "Detected Category": "N/A"})
            st.session_state.preview_data = prev
        if st.session_state.preview_data:
            with st.expander("🔍 Pre-processing Preview (Click to view)", expanded=False): st.table(pd.DataFrame(st.session_state.preview_data))

    if st.session_state.processed_data:
        st.divider(); render_overall_summary(st.session_state.processed_data, category="ALL")

else:
    cat_key = "AB18" if "AB18" in page else "AB17" if "AB17" in page else "AB16"
    st.title(page)
    if not st.session_state.processed_data: st.warning("Please upload and process files first.")
    else:
        res = [b for b in st.session_state.processed_data if b.get("Category", "").startswith(cat_key)]
        if not res: st.info(f"No {cat_key} items found.")
        else:
            missing_bidder_entries = [b for b in res if not b.get("Bidder")]
            visible_res = [b for b in res if b.get("Bidder")]
            if missing_bidder_entries:
                st.warning(build_missing_bidder_warning(missing_bidder_entries))
            if not visible_res:
                pass
            else:
                c_bidders = sorted(list(set(b.get("Bidder") for b in visible_res)))
                tab_names = ["📊 Summary"] + [b for b in all_processed_bidders if b in c_bidders]
                tabs = st.tabs(tab_names)
                with tabs[0]: render_overall_summary(visible_res, category=cat_key)
                for i, b_name in enumerate(tab_names[1:]):
                    with tabs[i+1]:
                        bidder_entries = [b for b in visible_res if b["Bidder"] == b_name]
                        if cat_key == "AB16":
                            merged_entry = merge_bidder_entries(bidder_entries, merged_category="AB16")
                            if merged_entry:
                                render_bidder_analysis(merged_entry)
                        else:
                            for entry in bidder_entries: render_bidder_analysis(entry); st.markdown("---")

if st.session_state.processed_data:
    st.sidebar.divider()
    if st.sidebar.button("🗑️ Reset Investigation"): st.session_state.processed_data = []; st.session_state.preview_data = []; st.rerun()
