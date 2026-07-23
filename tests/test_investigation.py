import io
import pytest
import openpyxl
import pandas as pd
from fastapi.testclient import TestClient

# Import components from our project
from backend.main import app
import backend.main
from backend.core.fuzzy_utils import robust_clean, fuzzy_match_name
from backend.services.analyzer import identify_category
from backend.processors.ab18_processor import AB18Processor
from backend.processors.ab16_processor import AB16Processor
from backend.thriller.pd_analyzer import PDAnalyzer


# =====================================================================
# 1. UNIT TESTS FOR UTILITIES
# =====================================================================

def test_robust_clean():
    """Test string cleaning utility used for bidder and manufacturer normalization."""
    # Test removing corporate suffixes
    assert robust_clean("Thai Yazaki Co., Ltd.") == "THAI YAZAKI"
    assert robust_clean("Phelps Dodge Thailand") == "PHELPS DODGE"
    
    # Test removing consortium and joint venture terms
    assert robust_clean("The Consortium of Demco Plc.") == "DEMCO"
    assert robust_clean("Joint Venture of A and B Ltd.") == "A AND B"
    
    # Test character cleaning and normalization
    assert robust_clean("A.B. C-D & Co.") == "A B C D"
    assert robust_clean("NAN") == ""
    assert robust_clean("-") == ""


def test_fuzzy_match_name():
    """Test matching names with high fuzzy similarity to canonical master names."""
    master_mfrs = {
        "THAI YAZAKI": "Thai Yazaki Co., Ltd.",
        "PHELPS DODGE": "Phelps Dodge (Thailand) Co., Ltd.",
        "SIEMENS": "Siemens AG"
    }
    
    # Exact match after cleaning
    assert fuzzy_match_name("Thai Yazaki Co., Ltd.", master_mfrs) == "Thai Yazaki Co., Ltd."
    
    # Fuzzy match with spelling mistake
    assert fuzzy_match_name("Thai Yazaky", master_mfrs) == "Thai Yazaki Co., Ltd."
    assert fuzzy_match_name("Pheleps Dodge", master_mfrs) == "Phelps Dodge (Thailand) Co., Ltd."
    
    # No match below threshold - should return original input
    assert fuzzy_match_name("Unknown Mfr", master_mfrs) == "Unknown Mfr"


def test_identify_category():
    """Test that sheets are correctly categorized based on key A1/A2 cells."""
    wb = openpyxl.Workbook()
    
    # Check AB18
    ws1 = wb.create_sheet("AB18_Sheet")
    ws1["A2"] = "LOW VOLTAGE CABLE AND CONDUCTOR"
    assert identify_category(ws1) == "AB18"
    
    # Check AB16 Terminations
    ws2 = wb.create_sheet("AB16_Term_Sheet")
    ws2["A1"] = "Cable terminations for XLPE"
    assert identify_category(ws2) == "AB16_TERM"
    
    # Check AB16 Cleats
    ws3 = wb.create_sheet("AB16_Cleat_Sheet")
    ws3["A1"] = "Cable Cleats Specification"
    assert identify_category(ws3) == "AB16_CLEAT"
    
    # Check AB17
    ws4 = wb.create_sheet("AB17_Sheet")
    ws4["A1"] = "XLPE Power cable proposal data"
    assert identify_category(ws4) == "AB17"
    
    # Unknown sheet
    ws5 = wb.create_sheet("Random_Sheet")
    ws5["A1"] = "Summary of prices"
    assert identify_category(ws5) is None


def test_ab17_e_section_flags_missing_copper_when_text_only_contains_calculated():
    """Ensure 'Calculated' does not falsely satisfy the Cu material check."""
    analyzer = PDAnalyzer("missing-standards-dir")
    df = pd.DataFrame([
        ["e.", "Calculated area of conductor", "10"],
    ])

    results = analyzer.analyze_sections(df)

    assert results["e"]["status"] == "ผิด"
    assert "Copper" in results["e"]["comment"]


def test_ab17_e_section_accepts_cu_as_standalone_material_token():
    """Allow valid copper shorthand like 'Cu conductor'."""
    analyzer = PDAnalyzer("missing-standards-dir")
    df = pd.DataFrame([
        ["e.", "Conductor material: Cu", "10"],
    ])

    results = analyzer.analyze_sections(df)

    assert "Copper" not in results["e"]["comment"]


# =====================================================================
# 2. INTEGRATION TESTS FOR PROCESSORS (USING IN-MEMORY WORKBOOKS)
# =====================================================================

def create_mock_ab18_worksheet(bidder="Thai Contractor", schedule="1AB18", 
                               item_no="1AB18-001", mfr="Thai Yazaki", 
                               country="TH", standard="conforming with EGAT requirement (TIS11-2559)",
                               ss=""):
    """Helper to build an openpyxl worksheet conforming to AB18 layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1AB18"
    
    # Set headers and metadata
    ws["A2"] = "LOW VOLTAGE CABLE AND CONDUCTOR"
    ws["C11"] = bidder
    ws["F8"] = "RTS3-S-02"
    ws["L8"] = schedule
    ws["L11"] = item_no
    
    # Set Power Cable fields
    ws["H15"] = mfr
    ws["H16"] = country
    ws["H18"] = standard
    ws["H20"] = ss
    
    return ws


def test_ab18_processor_pass():
    """Verify standard pass scenario: correct metadata, correct standards, and approved mfr."""
    mock_mfrs = {"THAI YAZAKI": "Thai Yazaki Co., Ltd."}
    mock_pools = (["Thai Yazaki Co., Ltd."], [])  # (General, Overhead)
    
    processor = AB18Processor(mock_mfrs, mock_pools)
    ws = create_mock_ab18_worksheet(
        bidder="Thai Contractor", 
        schedule="1AB18", 
        item_no="1AB18-001",
        mfr="Thai Yazaki", 
        country="TH"
    )
    
    # Mock price schedule mapping
    master_equip = {"1AB18-001": "1AB18-001 Power Cable 3x185 sq.mm"}
    equip_order = ["1AB18-001 Power Cable 3x185 sq.mm"]
    
    res = processor.process_sheet(ws, "proposal.xlsx", master_equip, equip_order=equip_order)
    
    assert res != {}
    assert res["BidderRaw"] == "Thai Contractor"
    assert res["Schedule"] == "1AB18"
    assert res["Status"] == "VALID"
    assert len(res["Issues"]) == 0
    
    eq_data = res["Equipment"]["1AB18-001 Power Cable 3x185 sq.mm"][0]
    assert eq_data["manufacturer"] == "Thai Yazaki Co., Ltd."
    assert eq_data["country"] == "THAILAND"
    assert eq_data["result"] == "Pass"
    assert eq_data["comment"] == ""


def test_ab18_processor_fail_standard():
    """Verify processor flags standard mismatch."""
    mock_mfrs = {"THAI YAZAKI": "Thai Yazaki Co., Ltd."}
    mock_pools = (["Thai Yazaki Co., Ltd."], [])
    
    processor = AB18Processor(mock_mfrs, mock_pools)
    
    # Incorrect standard 'TIS11-2222' (Required is TIS11-2559)
    ws = create_mock_ab18_worksheet(
        mfr="Thai Yazaki", 
        standard="conforming with EGAT requirement (TIS11-2222)"
    )
    
    master_equip = {"1AB18-001": "1AB18-001 Power Cable 3x185 sq.mm"}
    equip_order = ["1AB18-001 Power Cable 3x185 sq.mm"]
    
    res = processor.process_sheet(ws, "proposal.xlsx", master_equip, equip_order=equip_order)
    
    assert res["Status"] == "INVALID"
    assert any("Standard mismatch" in issue for issue in res["Issues"])
    
    eq_data = res["Equipment"]["1AB18-001 Power Cable 3x185 sq.mm"][0]
    assert eq_data["result"] == "Fail"
    assert "Standard mismatch" in eq_data["comment"]


def test_ab18_processor_fail_unapproved_mfr():
    """Verify processor flags manufacturer not in EGAT Approved Pool."""
    mock_mfrs = {"UNAPPROVED WIRE": "Unapproved Wire Corp"}
    mock_pools = (["Thai Yazaki Co., Ltd."], []) # Unapproved Wire not in General Pool
    
    processor = AB18Processor(mock_mfrs, mock_pools)
    ws = create_mock_ab18_worksheet(mfr="Unapproved Wire")
    
    master_equip = {"1AB18-001": "1AB18-001 Power Cable 3x185 sq.mm"}
    equip_order = ["1AB18-001 Power Cable 3x185 sq.mm"]
    
    res = processor.process_sheet(ws, "proposal.xlsx", master_equip, equip_order=equip_order)
    
    eq_data = res["Equipment"]["1AB18-001 Power Cable 3x185 sq.mm"][0]
    assert eq_data["result"] == "Fail"
    assert "Manufacturer not in EGAT Pool" in eq_data["comment"]


# =====================================================================
# 3. ENDPOINT INTEGRATION TESTS
# =====================================================================

def test_api_upload_flow():
    """Test the complete API upload flow using FastAPI TestClient with mock files."""
    # 1. Create a dummy price schedule Excel
    ps_wb = openpyxl.Workbook()
    ps_ws = ps_wb.active
    ps_ws.title = "Price Schedule"
    # Row 4, col 9 (I4) must identify category
    ps_ws["I4"] = "LOW VOLTAGE CABLE AND CONDUCTOR"
    # Row 16, cols I and J (maps to pandas row index 14)
    ps_ws.cell(row=16, column=9, value="1AB18-001") # Column I
    ps_ws.cell(row=16, column=10, value="Power Cable 3x185 sq.mm") # Column J
    
    ps_stream = io.BytesIO()
    ps_wb.save(ps_stream)
    ps_bytes = ps_stream.getvalue()

    # 2. Create a dummy proposal Excel
    prop_wb = openpyxl.Workbook()
    prop_ws = prop_wb.active
    prop_ws.title = "Proposal Sheet"
    prop_ws["A2"] = "LOW VOLTAGE CABLE AND CONDUCTOR"
    prop_ws["C11"] = "TE CONTRACTOR"
    prop_ws["F8"] = "RTS3-S-02"
    prop_ws["L8"] = "1AB18"
    prop_ws["L11"] = "1AB18-001"
    prop_ws["H15"] = "Thai Yazaki"
    prop_ws["H16"] = "TH"
    prop_ws["H18"] = "conforming with EGAT requirement (TIS11-2559)"
    
    prop_stream = io.BytesIO()
    prop_wb.save(prop_stream)
    prop_bytes = prop_stream.getvalue()

    # 3. Set Mock Globals on main module to isolate from real file system
    backend.main.MASTER_MANUFACTURERS = {"THAI YAZAKI": "Thai Yazaki Co., Ltd."}
    backend.main.EGAT_POOLS = (["Thai Yazaki Co., Ltd."], [])
    
    # Re-initialize AB18 processor with mock data
    backend.main.PROCESSORS["AB18"] = AB18Processor(
        backend.main.MASTER_MANUFACTURERS,
        backend.main.EGAT_POOLS
    )

    client = TestClient(app)
    
    # Call endpoint
    response = client.post(
        "/upload",
        files=[
            ("files", ("proposal_1.xlsx", prop_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("price_schedules", ("price_schedule.xlsx", ps_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
        ]
    )
    
    # Verify API Response
    assert response.status_code == 200
    json_data = response.json()
    assert "data" in json_data
    assert len(json_data["data"]) > 0
    
    bidder_entry = json_data["data"][0]
    assert bidder_entry["Bidder"] == "TE CONTRACTOR"
    assert bidder_entry["Category"] == "AB18"
    assert bidder_entry["Status"] == "VALID"
    assert len(bidder_entry["Issues"]) == 0
    
    # Verify parsed equipment structure
    equipment_map = bidder_entry["Equipment"]
    item_key = "1AB18-001 Power Cable 3x185 sq.mm"
    assert item_key in equipment_map
    assert equipment_map[item_key][0]["manufacturer"] == "Thai Yazaki Co., Ltd."
    assert equipment_map[item_key][0]["result"] == "Pass"
