import os
import re
import shutil
import tempfile
import pandas as pd
from ..processors.base import BaseProcessor
from typing import List, Dict, Any
from .pd_analyzer import PDAnalyzer
from ..shared.proposal_helpers import extract_item_id_from_pd, extract_detailed_manufacturer_from_pd
from ..shared.standard_paths import category_standard_path, category_standards_dir
from ..core.excel_utils import get_cell_val
from ..config import COMMON_CELLS
from openpyxl import load_workbook

# Initialize PDAnalyzer with the standards directory
STANDARDS_DIR = category_standards_dir("ab17", fallback_to_legacy=True)
pd_analyzer = PDAnalyzer(STANDARDS_DIR)

AB17_D_ONLY_ITEM = "1AB17-1"
AB17_A_ALLOWED_MFR_LE_33 = [
    "BANGKOK CABLE CO., LTD.",
    "THAI YAZAKI ELECTRIC WIRE CO., LTD.",
    "CHAROONG THAI WIRE & CABLE PUBLIC CO., LTD.",
    "PHELPS DODGE INTERNATIONAL THAILAND LTD."
]
AB17_A_ALLOWED_MFR_EQ_115 = list(AB17_A_ALLOWED_MFR_LE_33)
AB17_C_STD_LE_35 = "conforming with EGAT requirement ( ICEA Pub. No S-93-639)"
AB17_C_STD_60_TO_115 = "conforming with EGAT requirement (TIS 2202)"
AB17_C_STD_EQ_230 = "conforming with EGAT requirement (IEC 62067)"
AB17_K_EXPECTED_LE_33 = "copper tape"
AB17_K_EXPECTED_EQ_69_115 = "copper wire"
AB17_K_EXPECTED_EQ_230 = "metal aluminum corrugated shield"
AB17_Q_MIN_VALUE = 10.0
AB17_O_FIXED_BIL_MAP = {
    15.0: 110.0,
    25.0: 150.0,
    35.0: 200.0
}
AB17_O_TABLE_115_FILENAME = "o. Impulse Withstand Voltage (BIL) 115kV above.xlsx"
AB17_O_TABLE_EHV_FILENAME = "o. Impulse Withstand Voltage (BIL) EHV.xlsx"
AB17_O_TABLE_CACHE: Dict[str, List[Dict[str, float]]] = {}
AB17_R_TABLE_FILENAME = "r. Max Resistance at 20 °C of Conductor .xlsx"
AB17_R_TABLE_CACHE: List[Dict[str, float]] = []
AB17_O_TABLE_115_FALLBACK_ROWS = [
    {"min_kv": 45.0, "max_kv": 47.0, "bil_kv": 250.0},
    {"min_kv": 60.0, "max_kv": 69.0, "bil_kv": 325.0},
    {"min_kv": 110.0, "max_kv": 115.0, "bil_kv": 550.0},
    {"min_kv": 132.0, "max_kv": 138.0, "bil_kv": 650.0},
    {"min_kv": 150.0, "max_kv": 161.0, "bil_kv": 750.0}
]
AB17_O_TABLE_EHV_FALLBACK_ROWS = [
    {"min_kv": 220.0, "max_kv": 230.0, "bil_kv": 1050.0},
    {"min_kv": 275.0, "max_kv": 287.0, "bil_kv": 1050.0},
    {"min_kv": 330.0, "max_kv": 345.0, "bil_kv": 1175.0},
    {"min_kv": 380.0, "max_kv": 400.0, "bil_kv": 1425.0},
    {"min_kv": 500.0, "max_kv": 500.0, "bil_kv": 1550.0}
]
AB17_R_FALLBACK_ROWS = [
    {"area": 0.5, "max_resistance": 36.0},
    {"area": 0.75, "max_resistance": 24.5},
    {"area": 1.0, "max_resistance": 18.1},
    {"area": 1.5, "max_resistance": 12.1},
    {"area": 2.5, "max_resistance": 7.41},
    {"area": 4.0, "max_resistance": 4.61},
    {"area": 6.0, "max_resistance": 3.08},
    {"area": 10.0, "max_resistance": 1.83},
    {"area": 16.0, "max_resistance": 1.15},
    {"area": 25.0, "max_resistance": 0.727},
    {"area": 35.0, "max_resistance": 0.524},
    {"area": 50.0, "max_resistance": 0.387},
    {"area": 70.0, "max_resistance": 0.268},
    {"area": 95.0, "max_resistance": 0.193},
    {"area": 120.0, "max_resistance": 0.153},
    {"area": 150.0, "max_resistance": 0.124},
    {"area": 185.0, "max_resistance": 0.0991},
    {"area": 240.0, "max_resistance": 0.0754},
    {"area": 300.0, "max_resistance": 0.0601},
    {"area": 400.0, "max_resistance": 0.0470},
    {"area": 500.0, "max_resistance": 0.0366},
    {"area": 630.0, "max_resistance": 0.0283},
    {"area": 800.0, "max_resistance": 0.0221},
    {"area": 1000.0, "max_resistance": 0.0176},
    {"area": 1200.0, "max_resistance": 0.0151},
    {"area": 1400.0, "max_resistance": 0.0129},
    {"area": 1600.0, "max_resistance": 0.0113},
    {"area": 1800.0, "max_resistance": 0.0101},
    {"area": 2000.0, "max_resistance": 0.0090},
    {"area": 2500.0, "max_resistance": 0.0072},
    {"area": 3000.0, "max_resistance": 0.0060},
    {"area": 3200.0, "max_resistance": 0.0056},
    {"area": 3500.0, "max_resistance": 0.0051}
]

def _extract_section_row_map(df_pd: pd.DataFrame) -> Dict[str, int]:
    row_map = {}
    for idx, row in df_pd.iterrows():
        row_str = " ".join(str(v) for v in row.values if pd.notna(v)).strip()
        match = re.search(r"^\s*([a-s])[\.\)]", row_str, re.I)
        if match:
            row_map[match.group(1).lower()] = idx
    return row_map

def _extract_first_meaningful_number(text: str) -> float:
    if not text:
        return 0.0
    for match in re.finditer(r"\d+(?:\.\d+)?", str(text)):
        value = float(match.group(0))
        if value not in [60228, 60502, 2024, 2025, 2026, 2027]:
            return value
    return 0.0

def _parse_localized_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    text = re.sub(r"(?<=\d)\s+(?=\d)", "", text)
    text = text.replace(",", ".")
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return 0.0

    parsed_value = float(match.group(0))
    if parsed_value in [60228, 60502, 2024, 2025, 2026, 2027]:
        return 0.0
    return parsed_value

def _normalize_company_name(name: str) -> str:
    normalized = str(name or "").upper()
    normalized = normalized.replace("CHAROOG", "CHAROONG")
    normalized = re.sub(r"[^A-Z0-9]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized

def _normalize_standard_text(text: str) -> str:
    normalized = str(text or "").upper()
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized

def _normalize_keyword_text(text: str) -> str:
    normalized = str(text or "").lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized

def _is_blank_text(text: str) -> bool:
    cleaned = str(text or "").strip()
    return not cleaned or cleaned.lower() == "nan"

def _extract_section_text(df_pd: pd.DataFrame, key: str, row_limit: int = 3) -> str:
    row_map = _extract_section_row_map(df_pd)
    start_idx = row_map.get(str(key or "").lower())
    if start_idx is None:
        return ""

    collected_cells = []
    for offset in range(row_limit):
        curr_idx = start_idx + offset
        if curr_idx >= len(df_pd):
            break

        row_data = df_pd.iloc[curr_idx]
        for col_idx, value in enumerate(row_data.values):
            if pd.isna(value):
                continue
            cell_text = str(value).strip()
            if not cell_text:
                continue
            if offset == 0 and col_idx == 0:
                continue
            collected_cells.append(cell_text)

    return " ".join(collected_cells)

def _extract_section_numeric_values(df_pd: pd.DataFrame, key: str, row_limit: int = 3, max_values: int = 2) -> List[float]:
    row_map = _extract_section_row_map(df_pd)
    start_idx = row_map.get(str(key or "").lower())
    if start_idx is None:
        return []

    found_values = []
    for offset in range(row_limit):
        curr_idx = start_idx + offset
        if curr_idx >= len(df_pd):
            break

        row_data = df_pd.iloc[curr_idx]
        for col_idx, value in enumerate(row_data.values):
            if pd.isna(value):
                continue
            if offset == 0 and col_idx == 0:
                continue

            if isinstance(value, (int, float)):
                numeric_value = float(value)
                if numeric_value not in [60228, 60502, 2024, 2025, 2026, 2027]:
                    found_values.append(numeric_value)
            else:
                for match in re.finditer(r"\d+(?:\.\d+)?", str(value)):
                    numeric_value = float(match.group(0))
                    if numeric_value not in [60228, 60502, 2024, 2025, 2026, 2027]:
                        found_values.append(numeric_value)

            if len(found_values) >= max_values:
                return found_values[:max_values]

    return found_values[:max_values]

def _extract_section_localized_numeric_values(
    df_pd: pd.DataFrame,
    key: str,
    row_limit: int = 3,
    max_values: int = 2
) -> List[float]:
    row_map = _extract_section_row_map(df_pd)
    start_idx = row_map.get(str(key or "").lower())
    if start_idx is None:
        return []

    found_values = []
    for offset in range(row_limit):
        curr_idx = start_idx + offset
        if curr_idx >= len(df_pd):
            break

        row_data = df_pd.iloc[curr_idx]
        for col_idx, value in enumerate(row_data.values):
            if pd.isna(value):
                continue
            if offset == 0 and col_idx == 0:
                continue

            parsed_value = _parse_localized_float(value)
            if parsed_value > 0:
                found_values.append(parsed_value)

            if len(found_values) >= max_values:
                return found_values[:max_values]

    return found_values[:max_values]

def _extract_d_voltage_value(df_pd: pd.DataFrame) -> float:
    row_map = _extract_section_row_map(df_pd)
    start_idx = row_map.get("d")
    if start_idx is None:
        return 0.0

    for offset in range(4):
        curr_idx = start_idx + offset
        if curr_idx >= len(df_pd):
            break

        row_text = " ".join(str(v) for v in df_pd.iloc[curr_idx].values if pd.notna(v)).strip()
        kv_match = re.search(r"(\d+(?:\.\d+)?)\s*kv", row_text, re.I)
        if kv_match:
            return float(kv_match.group(1))

        if "kv" in row_text.lower():
            kv_values = [float(match.group(0)) for match in re.finditer(r"\d+(?:\.\d+)?", row_text)]
            kv_values = [value for value in kv_values if value not in [60228, 60502, 2024, 2025, 2026, 2027]]
            if kv_values:
                return max(kv_values)

        v_match = re.search(r"(\d{2,})\s*v\b", row_text, re.I)
        if v_match:
            raw_value = float(v_match.group(1))
            return raw_value / 1000 if raw_value >= 1000 else raw_value

        fallback_value = _extract_first_meaningful_number(row_text)
        if fallback_value > 0:
            return fallback_value

    return 0.0

def _parse_voltage_range_cell(text: str) -> Dict[str, float]:
    cleaned = str(text or "").strip()
    if not cleaned:
        return {}

    range_match = re.search(r"(\d+(?:\.\d+)?)\s*to\s*(\d+(?:\.\d+)?)", cleaned, re.I)
    if range_match:
        return {
            "min_kv": float(range_match.group(1)),
            "max_kv": float(range_match.group(2))
        }

    exact_match = re.search(r"^\s*(\d+(?:\.\d+)?)\s*$", cleaned)
    if exact_match:
        exact_value = float(exact_match.group(1))
        return {
            "min_kv": exact_value,
            "max_kv": exact_value
        }

    return {}

def _load_ab17_o_table_rows(filename: str) -> List[Dict[str, float]]:
    cached_rows = AB17_O_TABLE_CACHE.get(filename)
    if cached_rows is not None:
        return cached_rows

    fallback_rows = []
    if filename == AB17_O_TABLE_115_FILENAME:
        fallback_rows = list(AB17_O_TABLE_115_FALLBACK_ROWS)
    elif filename == AB17_O_TABLE_EHV_FILENAME:
        fallback_rows = list(AB17_O_TABLE_EHV_FALLBACK_ROWS)

    table_path = category_standard_path("ab17", filename, subdir="for 1AB17")
    temp_copy_path = None
    workbook = None

    try:
        workbook = load_workbook(table_path, data_only=True, read_only=True)
    except PermissionError:
        try:
            fd, temp_copy_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            shutil.copyfile(table_path, temp_copy_path)
            workbook = load_workbook(temp_copy_path, data_only=True, read_only=True)
        except PermissionError:
            AB17_O_TABLE_CACHE[filename] = fallback_rows
            return fallback_rows

    parsed_rows: List[Dict[str, float]] = []
    try:
        ws = workbook.active
        for row in ws.iter_rows(values_only=True):
            first_cell = row[0] if row else None
            range_info = _parse_voltage_range_cell(first_cell)
            if not range_info:
                continue

            bil_value = 0.0
            for cell_value in reversed(row[1:]):
                if cell_value is None:
                    continue
                cell_text = str(cell_value).strip()
                if not cell_text:
                    continue
                cell_numbers = [
                    float(match.group(0))
                    for match in re.finditer(r"\d+(?:\.\d+)?", cell_text)
                ]
                if cell_numbers:
                    bil_value = cell_numbers[-1]
                    break

            if bil_value > 0:
                parsed_rows.append({
                    "min_kv": range_info["min_kv"],
                    "max_kv": range_info["max_kv"],
                    "bil_kv": bil_value
                })
    finally:
        if workbook is not None:
            workbook.close()
        if temp_copy_path and os.path.exists(temp_copy_path):
            os.remove(temp_copy_path)

    final_rows = parsed_rows or fallback_rows
    AB17_O_TABLE_CACHE[filename] = final_rows
    return final_rows

def _load_ab17_r_table_rows() -> List[Dict[str, float]]:
    global AB17_R_TABLE_CACHE
    if AB17_R_TABLE_CACHE:
        return AB17_R_TABLE_CACHE

    table_path = category_standard_path("ab17", AB17_R_TABLE_FILENAME, subdir="for 1AB17")
    temp_copy_path = None
    workbook = None

    try:
        workbook = load_workbook(table_path, data_only=True, read_only=True)
    except PermissionError:
        try:
            fd, temp_copy_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            shutil.copyfile(table_path, temp_copy_path)
            workbook = load_workbook(temp_copy_path, data_only=True, read_only=True)
        except PermissionError:
            AB17_R_TABLE_CACHE = list(AB17_R_FALLBACK_ROWS)
            return AB17_R_TABLE_CACHE

    parsed_rows: List[Dict[str, float]] = []
    try:
        ws = workbook.active
        for row in ws.iter_rows(values_only=True):
            area_value = _parse_localized_float(row[0] if row else None)
            resistance_value = _parse_localized_float(row[7] if row and len(row) > 7 else None)
            if area_value > 0 and resistance_value > 0:
                parsed_rows.append({
                    "area": area_value,
                    "max_resistance": resistance_value
                })
    finally:
        if workbook is not None:
            workbook.close()
        if temp_copy_path and os.path.exists(temp_copy_path):
            os.remove(temp_copy_path)

    AB17_R_TABLE_CACHE = parsed_rows or list(AB17_R_FALLBACK_ROWS)
    return AB17_R_TABLE_CACHE

def _lookup_ab17_r_expected_max_resistance(area_value: float) -> Dict[str, Any]:
    if area_value <= 0:
        return {
            "applied": False,
            "expected_value": 0.0
        }

    tolerance = max(0.25, area_value * 0.01)
    for row in _load_ab17_r_table_rows():
        std_area = float(row.get("area", 0.0) or 0.0)
        if abs(std_area - area_value) <= tolerance:
            return {
                "applied": True,
                "expected_value": float(row.get("max_resistance", 0.0) or 0.0),
                "matched_area": std_area
            }

    return {
        "applied": True,
        "expected_value": 0.0,
        "lookup_failed": True
    }

def _lookup_ab17_o_expected_bil(d_voltage_kv: float) -> Dict[str, Any]:
    tolerance = 0.25

    for fixed_voltage, fixed_bil in AB17_O_FIXED_BIL_MAP.items():
        if abs(d_voltage_kv - fixed_voltage) <= tolerance:
            return {
                "applied": True,
                "expected_value": fixed_bil,
                "source_label": f"ค่าคงที่สำหรับ {fixed_voltage:g} kV"
            }

    if (115.0 - tolerance) <= d_voltage_kv <= (161.0 + tolerance):
        table_filename = AB17_O_TABLE_115_FILENAME
        source_label = "ตาราง o. Impulse Withstand Voltage (BIL) 115kV above"
    elif d_voltage_kv >= (220.0 - tolerance):
        table_filename = AB17_O_TABLE_EHV_FILENAME
        source_label = "ตาราง o. Impulse Withstand Voltage (BIL) EHV"
    else:
        return {
            "applied": False,
            "expected_value": 0.0,
            "source_label": ""
        }

    for row in _load_ab17_o_table_rows(table_filename):
        min_kv = row.get("min_kv", 0.0) - tolerance
        max_kv = row.get("max_kv", 0.0) + tolerance
        if min_kv <= d_voltage_kv <= max_kv:
            return {
                "applied": True,
                "expected_value": row.get("bil_kv", 0.0),
                "source_label": source_label,
                "table_filename": table_filename
            }

    return {
        "applied": True,
        "expected_value": 0.0,
        "source_label": source_label,
        "table_filename": table_filename,
        "lookup_failed": True
    }

def _extract_header_rule_text(ws) -> str:
    header_candidates = [ws.title]
    for row_idx in range(1, 7):
        row_values = []
        for col_idx in range(1, min(ws.max_column, 10) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            text = str(val).strip()
            if text:
                row_values.append(text)
        if row_values:
            header_candidates.append(" ".join(row_values))

    voltage_like_candidates = []
    header_keywords = [
        "kv", "voltage", "below", "above", "less", "more", "lower", "higher",
        "<", ">", "\u0e15\u0e48\u0e33\u0e01\u0e27\u0e48\u0e32", "\u0e2a\u0e39\u0e07\u0e01\u0e27\u0e48\u0e32",
        "\u0e44\u0e21\u0e48\u0e40\u0e01\u0e34\u0e19", "\u0e44\u0e21\u0e48\u0e19\u0e49\u0e2d\u0e22\u0e01\u0e27\u0e48\u0e32",
        "\u0e16\u0e36\u0e07"
    ]
    for candidate in header_candidates:
        lowered = candidate.lower()
        if re.search(r"\d", candidate) and any(keyword in lowered for keyword in header_keywords):
            voltage_like_candidates.append(candidate)

    if voltage_like_candidates:
        return " | ".join(voltage_like_candidates)

    return " | ".join(header_candidates)

def _parse_voltage_rule(header_text: str) -> Dict[str, Any]:
    normalized = str(header_text or "").strip()
    lowered = normalized.lower()
    values = [float(match.group(0)) for match in re.finditer(r"\d+(?:\.\d+)?", normalized)]
    values = [value for value in values if value not in [60228, 60502, 2024, 2025, 2026, 2027]]

    if not values:
        return {"mode": "unknown", "text": normalized}

    is_range = (
        len(values) >= 2 and (
            "between" in lowered or
            " to " in lowered or
            "\u0e16\u0e36\u0e07" in lowered or
            re.search(r"\d+(?:\.\d+)?\s*-\s*\d+(?:\.\d+)?", lowered)
        )
    )
    if is_range:
        return {
            "mode": "range",
            "text": normalized,
            "min_kv": min(values),
            "max_kv": max(values)
        }

    max_keywords = [
        "or less", "or below", "and below", "less than", "lower than", "up to",
        "<=", "\u2264", "not more than", "or lower", "\u0e15\u0e48\u0e33\u0e01\u0e27\u0e48\u0e32",
        "\u0e2b\u0e23\u0e37\u0e2d\u0e15\u0e48\u0e33\u0e01\u0e27\u0e48\u0e32", "\u0e44\u0e21\u0e48\u0e40\u0e01\u0e34\u0e19"
    ]
    min_keywords = [
        "or more", "or above", "and above", "greater than", "more than", "higher than",
        ">=", "\u2265", "not less than", "or higher", "\u0e21\u0e32\u0e01\u0e01\u0e27\u0e48\u0e32",
        "\u0e2b\u0e23\u0e37\u0e2d\u0e21\u0e32\u0e01\u0e01\u0e27\u0e48\u0e32", "\u0e2a\u0e39\u0e07\u0e01\u0e27\u0e48\u0e32",
        "\u0e44\u0e21\u0e48\u0e19\u0e49\u0e2d\u0e22\u0e01\u0e27\u0e48\u0e32"
    ]

    if any(keyword in lowered for keyword in max_keywords):
        return {"mode": "max", "text": normalized, "max_kv": max(values)}

    if any(keyword in lowered for keyword in min_keywords):
        return {"mode": "min", "text": normalized, "min_kv": min(values)}

    return {"mode": "exact", "text": normalized, "expected_kv": max(values)}

def _is_voltage_rule_match(d_voltage_kv: float, header_rule: Dict[str, Any]) -> bool:
    if d_voltage_kv <= 0:
        return False

    mode = header_rule.get("mode")
    tolerance = 0.25

    if mode == "max":
        return d_voltage_kv <= (header_rule.get("max_kv", 0.0) + tolerance)
    if mode == "min":
        return d_voltage_kv + tolerance >= header_rule.get("min_kv", 0.0)
    if mode == "range":
        min_kv = header_rule.get("min_kv", 0.0) - tolerance
        max_kv = header_rule.get("max_kv", 0.0) + tolerance
        return min_kv <= d_voltage_kv <= max_kv
    if mode == "exact":
        return abs(d_voltage_kv - header_rule.get("expected_kv", 0.0)) <= tolerance

    return False

def _evaluate_section_a_manufacturer_rule(manufacturer_name: str, d_voltage_kv: float) -> Dict[str, Any]:
    normalized_name = _normalize_company_name(manufacturer_name)
    tolerance = 0.25

    if d_voltage_kv > 0 and d_voltage_kv <= (33.0 + tolerance):
        allowed_names = AB17_A_ALLOWED_MFR_LE_33
        rule_label = "<= 33 kV"
    elif abs(d_voltage_kv - 115.0) <= tolerance:
        allowed_names = AB17_A_ALLOWED_MFR_EQ_115
        rule_label = "115 kV"
    else:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "allowed_names": []
        }

    normalized_allowed = {_normalize_company_name(name) for name in allowed_names}
    normalized_allowed.add(_normalize_company_name("CHAROONG THAI WIRE & CABLE PUBLIC CO., LTD."))
    if normalized_name in normalized_allowed:
        return {
            "applied": True,
            "passed": True,
            "comment": "",
            "allowed_names": allowed_names
        }

    return {
        "applied": True,
        "passed": False,
        "comment": (
            f"\u0e1c\u0e39\u0e49\u0e1c\u0e25\u0e34\u0e15 '{manufacturer_name or 'Unknown'}' "
            f"\u0e44\u0e21\u0e48\u0e2d\u0e22\u0e39\u0e48\u0e43\u0e19\u0e23\u0e32\u0e22\u0e0a\u0e37\u0e48\u0e2d\u0e17\u0e35\u0e48\u0e2d\u0e19\u0e38\u0e0d\u0e32\u0e15\u0e2a\u0e33\u0e2b\u0e23\u0e31\u0e1a {rule_label}. "
            f"\u0e1c\u0e39\u0e49\u0e1c\u0e25\u0e34\u0e15\u0e17\u0e35\u0e48\u0e22\u0e2d\u0e21\u0e23\u0e31\u0e1a: {', '.join(allowed_names)}"
        ),
        "allowed_names": allowed_names
    }

def _evaluate_section_c_standard_rule(item_no: str, standard_text: str, d_voltage_kv: float) -> Dict[str, Any]:
    normalized_item = str(item_no or "").strip().upper()
    if normalized_item != AB17_D_ONLY_ITEM:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_standard": ""
        }

    normalized_standard = _normalize_standard_text(standard_text)
    if "OTHER" in normalized_standard:
        return {
            "applied": True,
            "passed": False,
            "comment": "\u0e1c\u0e34\u0e14 \u0e01\u0e23\u0e38\u0e13\u0e32\u0e01\u0e23\u0e2d\u0e01 standard \u0e17\u0e35\u0e48\u0e43\u0e0a\u0e49\u0e07\u0e32\u0e19",
            "expected_standard": ""
        }

    tolerance = 0.25
    expected_standard = ""

    if d_voltage_kv > 0 and d_voltage_kv <= (35.0 + tolerance):
        expected_standard = AB17_C_STD_LE_35
    elif (60.0 - tolerance) <= d_voltage_kv <= (115.0 + tolerance):
        expected_standard = AB17_C_STD_60_TO_115
    elif abs(d_voltage_kv - 230.0) <= tolerance:
        expected_standard = AB17_C_STD_EQ_230
    else:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_standard": ""
        }

    normalized_expected = _normalize_standard_text(expected_standard)
    if normalized_standard == normalized_expected:
        return {
            "applied": True,
            "passed": True,
            "comment": "",
            "expected_standard": expected_standard
        }

    return {
        "applied": True,
        "passed": False,
        "comment": (
            f"\u0e40\u0e25\u0e37\u0e2d\u0e01 standard \u0e44\u0e21\u0e48\u0e16\u0e39\u0e01\u0e15\u0e49\u0e2d\u0e07\u0e2a\u0e33\u0e2b\u0e23\u0e31\u0e1a {d_voltage_kv:g} kV. "
            f"\u0e04\u0e27\u0e23\u0e40\u0e1b\u0e47\u0e19: {expected_standard}"
        ),
        "expected_standard": expected_standard
    }

def _evaluate_section_k_material_rule(item_no: str, k_text: str, d_voltage_kv: float) -> Dict[str, Any]:
    normalized_item = str(item_no or "").strip().upper()
    if normalized_item != AB17_D_ONLY_ITEM:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_material": ""
        }

    tolerance = 0.25
    expected_material = ""

    if d_voltage_kv > 0 and d_voltage_kv <= (33.0 + tolerance):
        expected_material = AB17_K_EXPECTED_LE_33
        required_terms = ["copper", "tape"]
    elif abs(d_voltage_kv - 69.0) <= tolerance or abs(d_voltage_kv - 115.0) <= tolerance:
        expected_material = AB17_K_EXPECTED_EQ_69_115
        required_terms = ["copper", "wire"]
    elif abs(d_voltage_kv - 230.0) <= tolerance:
        expected_material = AB17_K_EXPECTED_EQ_230
        required_terms = ["metal", "aluminum", "corrugated", "shield"]
    else:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_material": ""
        }

    normalized_k_text = _normalize_keyword_text(k_text)
    if all(term in normalized_k_text for term in required_terms):
        return {
            "applied": True,
            "passed": True,
            "comment": "",
            "expected_material": expected_material
        }

    return {
        "applied": True,
        "passed": False,
        "comment": (
            f"\u0e02\u0e49\u0e2d k \u0e40\u0e25\u0e37\u0e2d\u0e01\u0e27\u0e31\u0e2a\u0e14\u0e38\u0e44\u0e21\u0e48\u0e16\u0e39\u0e01\u0e15\u0e49\u0e2d\u0e07\u0e2a\u0e33\u0e2b\u0e23\u0e31\u0e1a {d_voltage_kv:g} kV. "
            f"\u0e04\u0e27\u0e23\u0e40\u0e1b\u0e47\u0e19: {expected_material}"
        ),
        "expected_material": expected_material
    }

def _evaluate_section_q_rule(item_no: str, q_values: List[float]) -> Dict[str, Any]:
    normalized_item = str(item_no or "").strip().upper()
    if normalized_item != AB17_D_ONLY_ITEM:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "values": q_values
        }

    labels = ["top", "bottom"]
    issues = []
    checked_values = q_values[:2]

    if not checked_values:
        return {
            "applied": True,
            "passed": False,
            "comment": (
                f"\u0e02\u0e49\u0e2d q \u0e44\u0e21\u0e48\u0e1e\u0e1a\u0e04\u0e48\u0e32\u0e17\u0e35\u0e48\u0e15\u0e49\u0e2d\u0e07\u0e15\u0e23\u0e27\u0e08 "
                f"\u0e42\u0e14\u0e22\u0e04\u0e48\u0e32\u0e14\u0e49\u0e32\u0e19\u0e1a\u0e19\u0e41\u0e25\u0e30\u0e14\u0e49\u0e32\u0e19\u0e25\u0e48\u0e32\u0e07\u0e15\u0e49\u0e2d\u0e07 >= {AB17_Q_MIN_VALUE:g}"
            ),
            "values": []
        }

    for idx, value in enumerate(checked_values):
        if value < AB17_Q_MIN_VALUE:
            if labels[idx] == "top":
                issues.append(
                    f"\u0e04\u0e48\u0e32\u0e0a\u0e48\u0e2d\u0e07\u0e1a\u0e19 {value:g} \u0e19\u0e49\u0e2d\u0e22\u0e01\u0e27\u0e48\u0e32 {AB17_Q_MIN_VALUE:g}"
                )
            else:
                issues.append(
                    f"\u0e04\u0e48\u0e32\u0e0a\u0e48\u0e2d\u0e07\u0e25\u0e48\u0e32\u0e07 {value:g} \u0e19\u0e49\u0e2d\u0e22\u0e01\u0e27\u0e48\u0e32 {AB17_Q_MIN_VALUE:g}"
                )

    if len(checked_values) < 2:
        issues.append("\u0e44\u0e21\u0e48\u0e1e\u0e1a\u0e04\u0e48\u0e32\u0e0a\u0e48\u0e2d\u0e07\u0e25\u0e48\u0e32\u0e07")

    if not issues:
        return {
            "applied": True,
            "passed": True,
            "comment": "",
            "values": checked_values
        }

    return {
        "applied": True,
        "passed": False,
        "comment": "\u0e02\u0e49\u0e2d q \u0e44\u0e21\u0e48\u0e1c\u0e48\u0e32\u0e19: " + "; ".join(issues),
        "values": checked_values
    }

def _evaluate_section_o_rule(item_no: str, o_values: List[float], d_voltage_kv: float) -> Dict[str, Any]:
    normalized_item = str(item_no or "").strip().upper()
    if normalized_item != AB17_D_ONLY_ITEM:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_value": 0.0,
            "values": o_values
        }

    if d_voltage_kv <= 0:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_value": 0.0,
            "values": o_values
        }

    lookup_result = _lookup_ab17_o_expected_bil(d_voltage_kv)
    if not lookup_result.get("applied"):
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_value": 0.0,
            "values": o_values
        }

    if lookup_result.get("lookup_failed"):
        return {
            "applied": True,
            "passed": False,
            "comment": (
                f"ข้อ o ไม่พบช่วงแรงดัน {d_voltage_kv:g} kV ใน {lookup_result.get('source_label', 'ตารางมาตรฐาน')}"
            ),
            "expected_value": 0.0,
            "values": o_values,
            "source_label": lookup_result.get("source_label", "")
        }

    expected_value = float(lookup_result.get("expected_value", 0.0) or 0.0)
    actual_value = float(o_values[0]) if o_values else 0.0
    tolerance = 0.25

    if actual_value <= 0:
        return {
            "applied": True,
            "passed": False,
            "comment": (
                f"ข้อ o ยังไม่ได้กรอกค่า Impulse Withstand Voltage (BIL) สำหรับ {d_voltage_kv:g} kV"
            ),
            "expected_value": expected_value,
            "values": o_values,
            "source_label": lookup_result.get("source_label", "")
        }

    if abs(actual_value - expected_value) <= tolerance:
        return {
            "applied": True,
            "passed": True,
            "comment": "",
            "expected_value": expected_value,
            "values": o_values,
            "source_label": lookup_result.get("source_label", "")
        }

    return {
        "applied": True,
        "passed": False,
        "comment": (
            f"ข้อ o กรอกค่า {actual_value:g} kV ไม่ถูกต้องสำหรับ {d_voltage_kv:g} kV. "
            f"ควรเป็น {expected_value:g} kV"
        ),
        "expected_value": expected_value,
        "values": o_values,
        "source_label": lookup_result.get("source_label", "")
    }

def _evaluate_section_r_rule(item_no: str, e_area_value: float, r_values: List[float]) -> Dict[str, Any]:
    normalized_item = str(item_no or "").strip().upper()
    # Temporarily skip section r validation for 1AB17-1 and leave it blank in output.
    if normalized_item == AB17_D_ONLY_ITEM:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_value": 0.0,
            "values": []
        }

    if normalized_item != AB17_D_ONLY_ITEM:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_value": 0.0,
            "values": r_values
        }

    if e_area_value <= 0:
        return {
            "applied": False,
            "passed": True,
            "comment": "",
            "expected_value": 0.0,
            "values": r_values
        }

    lookup_result = _lookup_ab17_r_expected_max_resistance(e_area_value)
    if lookup_result.get("lookup_failed"):
        return {
            "applied": True,
            "passed": False,
            "comment": (
                f"ข้อ r ไม่พบค่า Nominal cross-sectional area {e_area_value:g} mm² ในตาราง r"
            ),
            "expected_value": 0.0,
            "values": r_values
        }

    expected_value = float(lookup_result.get("expected_value", 0.0) or 0.0)
    actual_value = float(r_values[0]) if r_values else 0.0
    comparison_tolerance = max(0.0005, expected_value * 0.03)

    if actual_value <= 0:
        return {
            "applied": True,
            "passed": False,
            "comment": (
                f"ข้อ r ยังไม่ได้กรอกค่า Max Resistance at 20 °C of Conductor สำหรับพื้นที่หน้าตัด {e_area_value:g} mm²"
            ),
            "expected_value": expected_value,
            "values": r_values,
            "matched_area": lookup_result.get("matched_area", e_area_value)
        }

    if actual_value <= (expected_value + comparison_tolerance):
        return {
            "applied": True,
            "passed": True,
            "comment": "",
            "expected_value": expected_value,
            "values": r_values,
            "matched_area": lookup_result.get("matched_area", e_area_value)
        }

    return {
        "applied": True,
        "passed": False,
        "comment": (
            f"ข้อ r กรอกค่า {actual_value:g} Ω/km สูงเกินกว่าค่ามาตรฐานสำหรับพื้นที่หน้าตัด {e_area_value:g} mm². "
            f"ค่าที่ควรได้ต้องไม่เกิน {expected_value:g} Ω/km"
        ),
        "expected_value": expected_value,
        "values": r_values,
        "matched_area": lookup_result.get("matched_area", e_area_value)
    }

def _build_d_header_context(ws, df_pd: pd.DataFrame, item_no: str) -> Dict[str, Any]:
    normalized_item = str(item_no or "").strip().upper()
    if normalized_item != AB17_D_ONLY_ITEM:
        return {}

    d_voltage_kv = _extract_d_voltage_value(df_pd)
    header_text = _extract_header_rule_text(ws)
    header_rule = _parse_voltage_rule(header_text)
    is_match = _is_voltage_rule_match(d_voltage_kv, header_rule)

    if not is_match:
        return {}

    return {
        "item_no": normalized_item,
        "d_voltage_kv": d_voltage_kv,
        "header_rule": header_rule,
        "matched": True
    }

class AB17Processor(BaseProcessor):
    @staticmethod
    def _is_valid_bidder_candidate(value: Any) -> bool:
        text = str(value or "").strip()
        if not text:
            return False

        normalized = " ".join(text.upper().split())
        if "MANUFACTURER" in normalized or len(normalized) <= 2:
            return False

        # Some files place the schedule title/description in the bidder area.
        # Treat those as missing bidder instead of creating a fake bidder tab.
        title_like_markers = [
            "XLPE POWER CABLE",
            "TURNKEY SUPPLY",
            "KV AND BELOW",
            "[TURNKEY SUPPLY]",
        ]
        if any(marker in normalized for marker in title_like_markers):
            return False

        return True

    def get_category(self) -> str:
        return "AB17"

    def process_sheet(self, ws, filename: str, master_equip: Dict[str, str], **kwargs) -> Dict[str, Any]:
        """
        Processes an AB17 sheet using the integrated Thriller logic.
        """
        # 1. Convert to DataFrame for Thriller logic
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        df_pd = pd.DataFrame(data)

        # Metadata extraction
        bidder_raw = get_cell_val(ws, "C6")
        if not self._is_valid_bidder_candidate(bidder_raw):
            bidder_raw = ""

        if not bidder_raw:
            # If empty or contains label, search top rows for anything that looks like a name
            for r in range(5, 15):
                val = get_cell_val(ws, f"C{r}")
                if self._is_valid_bidder_candidate(val):
                    bidder_raw = val
                    break

        item_no = extract_item_id_from_pd(df_pd)
        if not bidder_raw:
            proc_ref = get_cell_val(ws, "E6")
            sched_no = get_cell_val(ws, "K6") or get_cell_val(ws, "L6") or get_cell_val(ws, "M6")
            return {
                "BidderRaw": "",
                "ProcRef": proc_ref,
                "Schedule": sched_no,
                "ItemNo": item_no or "Unknown Item",
                "Status": "INVALID",
                "Issues": [f"Forgot to fill bidder in file {filename}"],
                "Equipment": {}
            }

        mfr_data = extract_detailed_manufacturer_from_pd(df_pd)
        standard_text = get_cell_val(ws, "G16")
        section_e_values = _extract_section_localized_numeric_values(df_pd, "e", row_limit=4, max_values=1)
        section_k_text = _extract_section_text(df_pd, "k", row_limit=3)
        section_o_values = _extract_section_numeric_values(df_pd, "o", row_limit=3, max_values=1)
        section_q_values = _extract_section_numeric_values(df_pd, "q", row_limit=3, max_values=2)
        section_r_values = _extract_section_localized_numeric_values(df_pd, "r", row_limit=3, max_values=1)
        d_header_context = _build_d_header_context(ws, df_pd, item_no)
        d_voltage_kv = _extract_d_voltage_value(df_pd)
        e_area_value = float(section_e_values[0]) if section_e_values else 0.0
        section_a_rule = _evaluate_section_a_manufacturer_rule(mfr_data.get("manufacturer", ""), d_voltage_kv)
        section_c_rule = _evaluate_section_c_standard_rule(item_no, standard_text, d_voltage_kv)
        section_k_rule = _evaluate_section_k_material_rule(item_no, section_k_text, d_voltage_kv)
        section_o_rule = _evaluate_section_o_rule(item_no, section_o_values, d_voltage_kv)
        section_q_rule = _evaluate_section_q_rule(item_no, section_q_values)
        section_r_rule = _evaluate_section_r_rule(item_no, e_area_value, section_r_values)
        analysis_results = pd_analyzer.analyze_sections(df_pd)
        if section_k_rule.get("applied") and not section_k_rule.get("passed"):
            existing_k = analysis_results.get("k", {"status": "à¸œà¹ˆà¸²à¸™", "comment": ""})
            existing_comment = str(existing_k.get("comment", "")).strip()
            if existing_comment and existing_comment != "à¸œà¹ˆà¸²à¸™":
                merged_comment = f"{existing_comment}; {section_k_rule['comment']}"
            else:
                merged_comment = section_k_rule["comment"]
            analysis_results["k"] = {"status": "à¸œà¸´à¸”", "comment": merged_comment}

        # 2. Synthesize Results (Exact Thriller-style logic)
        section_names = {
            "e": "Area", "f": "Diameter", "g": "Cond. Screen", "h": "Insulation",
            "i": "Ins. Screen", "j": "OD Insulation", "k": "Metal Screen",
            "l": "Cushion", "m": "Oversheath", "n": "Overall Diameter"
        }
        
        error_sections = []
        detailed_comments = []
        custom_error_details = []
        custom_fix_suggestions = []
        missing_fields = []
        if str(item_no or "").strip().upper() == AB17_D_ONLY_ITEM:
            manufacturer_name = str(mfr_data.get("manufacturer", "") or "").strip()
            if not manufacturer_name or manufacturer_name.upper() == "UNKNOWN":
                missing_fields.append(("a. ผู้ผลิต", "a. ผู้ผลิต"))

            if d_voltage_kv <= 0:
                missing_fields.append((
                    "d. ยังไม่ได้กรอกค่าแรงดัน",
                    "d. กรุณากรอกค่าแรงดันในข้อ d ให้ชัดเจน"
                ))

            if _is_blank_text(standard_text):
                missing_fields.append(("c. มาตรฐาน", "c. มาตรฐาน"))

            if _is_blank_text(section_k_text):
                missing_fields.append(("k. วัสดุของ Metal Screen", "k. วัสดุของ Metal Screen"))

            if section_o_rule.get("applied") and len(section_o_values) == 0:
                missing_fields.append(("o. Impulse Withstand Voltage (BIL)", "o. Impulse Withstand Voltage (BIL)"))

            if e_area_value <= 0:
                missing_fields.append(("e. Nominal cross-sectional area", "e. Nominal cross-sectional area"))

            if section_r_rule.get("applied") and len(section_r_values) == 0:
                missing_fields.append(("r. Max Resistance at 20 °C of Conductor", "r. Max Resistance at 20 °C of Conductor"))

            if len(section_q_values) == 0:
                missing_fields.append(("q. ช่องบน", "q. ช่องบน"))
                missing_fields.append(("q. ช่องล่าง", "q. ช่องล่าง"))
            elif len(section_q_values) == 1:
                missing_fields.append(("q. ช่องล่าง", "q. ช่องล่าง"))

        if section_a_rule.get("applied") and not section_a_rule.get("passed"):
            error_sections.append("a. ผู้ผลิต")
            detailed_comments.append(f"a: {section_a_rule['comment']}")
            custom_error_details.append(
                f"a. ผู้ผลิต '{mfr_data.get('manufacturer', 'Unknown') or 'Unknown'}' ไม่อยู่ในเงื่อนไขที่กำหนดตามแรงดันในข้อ d"
            )
            custom_fix_suggestions.append(
                "a. กรุณาเลือกผู้ผลิตให้ตรงกับรายชื่อที่อนุญาตตามระดับแรงดันในข้อ d"
            )
        if section_c_rule.get("applied") and not section_c_rule.get("passed"):
            error_sections.append("c. มาตรฐาน")
            detailed_comments.append(f"c: {section_c_rule['comment']}")
            if "OTHER" in _normalize_standard_text(standard_text):
                custom_error_details.append("c. ยังไม่ได้ระบุ standard ที่ใช้งาน")
            else:
                custom_error_details.append(
                    f"c. standard ที่เลือกไม่ตรงกับเงื่อนไขสำหรับแรงดัน {d_voltage_kv:g} kV"
                )
            if "OTHER" in _normalize_standard_text(standard_text):
                custom_fix_suggestions.append("c. กรุณากรอก standard ที่ใช้งานให้ชัดเจน")
            elif section_c_rule.get("expected_standard"):
                custom_fix_suggestions.append(
                    f"c. กรุณาเลือก standard เป็น {section_c_rule['expected_standard']}"
                )
        if section_k_rule.get("applied") and not section_k_rule.get("passed"):
            custom_error_details.append(
                f"k. วัสดุของ Metal Screen ไม่ตรงกับเงื่อนไขสำหรับแรงดัน {d_voltage_kv:g} kV"
            )
            if section_k_rule.get("expected_material"):
                custom_fix_suggestions.append(
                    f"k. กรุณาระบุวัสดุเป็น {section_k_rule['expected_material']}"
                )
        if section_o_rule.get("applied") and not section_o_rule.get("passed"):
            error_sections.append("o. Impulse Withstand Voltage (BIL)")
            detailed_comments.append(f"o: {section_o_rule['comment']}")
            if len(section_o_values) == 0:
                custom_error_details.append("o. ยังไม่ได้กรอกค่า Impulse Withstand Voltage (BIL)")
            elif section_o_rule.get("lookup_failed"):
                custom_error_details.append(
                    f"o. ไม่พบช่วงแรงดัน {d_voltage_kv:g} kV ในตารางมาตรฐานที่ใช้อ้างอิง"
                )
            else:
                actual_o_value = float(section_o_values[0]) if section_o_values else 0.0
                custom_error_details.append(
                    f"o. ค่า Impulse Withstand Voltage (BIL) ที่กรอก ({actual_o_value:g} kV) ไม่ตรงกับเงื่อนไขสำหรับแรงดัน {d_voltage_kv:g} kV"
                )
            if section_o_rule.get("expected_value"):
                custom_fix_suggestions.append(
                    f"o. กรุณากรอกค่า Impulse Withstand Voltage (BIL) เป็น {section_o_rule['expected_value']:g} kV"
                )
        if section_q_rule.get("applied") and not section_q_rule.get("passed"):
            error_sections.append("q. ค่า")
            detailed_comments.append(f"q: {section_q_rule['comment']}")
            custom_error_details.append(f"q. {section_q_rule['comment']}")
            custom_fix_suggestions.append(
                f"q. กรุณากรอกค่าช่องบนและช่องล่างให้มากกว่าหรือเท่ากับ {AB17_Q_MIN_VALUE:g}"
            )
        if section_r_rule.get("applied") and not section_r_rule.get("passed"):
            error_sections.append("r. Max Resistance at 20 °C of Conductor")
            detailed_comments.append(f"r: {section_r_rule['comment']}")
            custom_error_details.append(f"r. {section_r_rule['comment']}")
            if section_r_rule.get("expected_value"):
                custom_fix_suggestions.append(
                    f"r. กรุณากรอกค่า Max Resistance at 20 °C of Conductor ไม่เกิน {section_r_rule['expected_value']:g} Ω/km ตามพื้นที่หน้าตัดจากข้อ e"
                )
        for sec_key in "efghijklmn":
            if sec_key in analysis_results:
                res = analysis_results[sec_key]
                if res["status"] == "ผิด":
                    name = section_names.get(sec_key, sec_key)
                    error_sections.append(f"❌ {sec_key}. {name}")
                    detailed_comments.append(f"• {sec_key}: {res['comment']}")
        
        if not error_sections:
            status = "VALID"
            final_status = "✅ ผ่านทุกหัวข้อ (e-n)"
            final_comment = "ตรวจสอบแล้วถูกต้องครบถ้วน"
        else:
            status = "INVALID"
            final_status = "\n".join(error_sections)
            final_comment = "พบจุดที่ต้องแก้ไข:\n" + "\n".join(detailed_comments)

        is_ab17_d_only_item = str(item_no or "").strip().upper() == AB17_D_ONLY_ITEM

        if custom_error_details or missing_fields:
            status_parts = []
            comment_parts = []

            if custom_error_details:
                status_parts.append(
                    "\u274c \u0e44\u0e21\u0e48\u0e1c\u0e48\u0e32\u0e19\u0e40\u0e07\u0e37\u0e48\u0e2d\u0e19\u0e44\u0e02:\n" +
                    "\n".join(custom_error_details)
                )
                comment_parts.append(
                    "\ud83d\udca1 \u0e02\u0e49\u0e2d\u0e40\u0e2a\u0e19\u0e2d\u0e41\u0e19\u0e30\u0e40\u0e1e\u0e37\u0e48\u0e2d\u0e41\u0e01\u0e49\u0e44\u0e02:\n" +
                    "\n".join(custom_fix_suggestions)
                )

            if missing_fields:
                missing_status_lines = [
                    "\u26a0\ufe0f \u0e25\u0e37\u0e21\u0e01\u0e23\u0e2d\u0e01:"
                ]
                missing_comment_lines = [
                    "\u26a0\ufe0f \u0e25\u0e37\u0e21\u0e01\u0e23\u0e2d\u0e01:"
                ]
                for short_name, detailed_name in missing_fields:
                    missing_status_lines.append(f"  {short_name}")
                    missing_comment_lines.append(f"  {detailed_name}")

                if status_parts and not is_ab17_d_only_item:
                    status_parts.append("\n" + "-" * 30)
                if comment_parts and not is_ab17_d_only_item:
                    comment_parts.append("\n" + "-" * 30)

                if not is_ab17_d_only_item:
                    status_parts.append("\n".join(missing_status_lines))
                comment_parts.append("\n".join(missing_comment_lines))

            if status == "INVALID" and not is_ab17_d_only_item:
                status_parts.append("\n" + "-" * 30)
                comment_parts.append("\n" + "-" * 30)
                status_parts.append(final_status)
                comment_parts.append(final_comment)

            final_status = "\n".join(status_parts)
            final_comment = "\n".join(comment_parts)
            status = "INVALID"

        # 3. Format result
        proc_ref = get_cell_val(ws, "E6")
        sched_no = get_cell_val(ws, "K6") or get_cell_val(ws, "L6") or get_cell_val(ws, "M6")
        
        equip_order = kwargs.get("equip_order", [])
        # Exact match with ID part to prevent picking AB16 items that mention AB17 in description
        eq_key = next((k for k in equip_order if item_no and k.split()[0].upper() == str(item_no).upper()), str(item_no) if item_no else "Unknown Item")
        display_result = "Pass" if status == "VALID" else "Fail"
        display_comment = "\n\n".join(part for part in [final_status, final_comment] if str(part).strip())

        return {
            "BidderRaw": bidder_raw,
            "ProcRef": proc_ref,
            "Schedule": sched_no,
            "ItemNo": item_no,
            "Status": status,
            "Issues": [final_comment] if status != "VALID" else [],
            "Equipment": {
                eq_key: [{
                    "manufacturer": mfr_data.get("manufacturer", "Unknown"),
                    "country": mfr_data.get("country", "") or "Unknown",
                    "type_model": mfr_data.get("type_model", ""),
                    "standard": standard_text,
                    "result": display_result,
                    "comment": display_comment,
                    "_ab17_d_header_context": d_header_context,
                    "_ab17_a_context": section_a_rule,
                    "_ab17_c_context": section_c_rule,
                    "_ab17_k_context": section_k_rule,
                    "_ab17_o_context": section_o_rule,
                    "_ab17_q_context": section_q_rule,
                    "_ab17_r_context": section_r_rule,
                    "source": f"{filename} > {ws.title}"
                }]
            }
        }

