import os
import re
import openpyxl
from io import BytesIO
from typing import List, Dict, Optional
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from datetime import datetime
from thefuzz import process as fuzz_process

from .config import PRICE_SCHEDULE_DEFAULT_PATH, UNWANTED_TERMS, BIDDER_MATCH_THRESHOLD, BIDDER_ALIASES
from .core.excel_utils import get_cell_val
from .core.fuzzy_utils import robust_clean
from .core.data_loader import load_master_manufacturers, load_egat_mfr_pools, extract_price_schedule_data
from .services.analyzer import identify_category
from .ab18 import AB18Processor
from .ab16 import AB16Processor
from .ab17 import AB17Processor

app = FastAPI(title="TodSaMong Overall API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,  # Set to False when using "*" for allow_origins
    allow_methods=["*"],
    allow_headers=["*"],
)

MASTER_MANUFACTURERS = load_master_manufacturers()
EGAT_POOLS = load_egat_mfr_pools()
PROCESSORS = {
    "AB18": AB18Processor(MASTER_MANUFACTURERS, EGAT_POOLS),
    "AB16_TERM": AB16Processor("AB16_TERM", MASTER_MANUFACTURERS),
    "AB16_CLEAT": AB16Processor("AB16_CLEAT", MASTER_MANUFACTURERS),
    "AB17": AB17Processor()
}

def is_valid_bidder(name: str) -> bool:
    if not name: return False
    n_up = str(name).upper().strip()
    invalid_keywords = [
        "COUNTRY OF ORIGIN", "MANUFACTURER", "SCHEDULE", "ITEM", 
        "SPECIFICATION", "TECHNICAL", "PROPOSAL DATA", "DESCRIPTION", 
        "GUARANTEE", "CONFORMING", "TOTAL PRICE", "SUMMARY", "TRANSPORTATION"
    ]
    title_like_markers = [
        "XLPE POWER CABLE",
        "TURNKEY SUPPLY",
        "KV AND BELOW",
    ]
    if any(kw in n_up for kw in invalid_keywords):
        return False
    if any(marker in n_up for marker in title_like_markers):
        return False
    if len(n_up) < 2:
        return False
    return True

@app.get("/")
async def root(): return {"status": "online"}

def clean_surrogates(val):
    if isinstance(val, str):
        return val.encode("utf-8", "ignore").decode("utf-8")
    elif isinstance(val, dict):
        return {clean_surrogates(k): clean_surrogates(v) for k, v in val.items()}
    elif isinstance(val, list):
        return [clean_surrogates(x) for x in val]
    return val

@app.post("/upload")
async def upload(files: List[UploadFile] = File(...), price_schedules: List[UploadFile] = File(None)):
    # 1. Load Master Schedule
    equip_order, master_equip = [], {}
    if price_schedules:
        for ps in price_schedules:
            content = await ps.read(); eo, me = extract_price_schedule_data(content, "ALL")
            equip_order.extend(eo); master_equip.update(me)
    elif os.path.exists(PRICE_SCHEDULE_DEFAULT_PATH):
        with open(PRICE_SCHEDULE_DEFAULT_PATH, "rb") as f:
            eo, me = extract_price_schedule_data(f.read(), "ALL"); equip_order.extend(eo); master_equip.update(me)
    
    if not equip_order: raise HTTPException(status_code=400, detail="Master Price Schedule missing or invalid.")

    # 2. Process Files
    raw_results, ab17_cable_data, all_contents = [], {}, []
    for f in files:
        content = await f.read(); all_contents.append((f.filename, content))
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            if "XLPE POWER" in str(get_cell_val(ws, "A1")).upper():
                it = ""
                for c in range(1, 15):
                    v = str(ws.cell(row=7, column=c).value or "").strip()
                    if re.match(r"\d+AB17-\d+", v, re.I): it = v.upper(); break
                if it:
                    ab17_cable_data[it] = {"min": get_cell_val(ws, "I38"), "max": get_cell_val(ws, "M38")}
                    try:
                        import pandas as pd
                        from .ab16.term_logic import extract_fields_from_pd
                        df_pd = pd.DataFrame(ws.values)
                        fields = extract_fields_from_pd(df_pd)
                        ab17_cable_data[it]["f_cond_od"] = fields.get("f_cond_od", "")
                        ab17_cable_data[it]["g_cond_screen"] = fields.get("g_cond_screen", "")
                        ab17_cable_data[it]["h_insulation"] = fields.get("h_insulation", "")
                    except Exception as e:
                        print(f"[DEBUG] Error extracting f,g,h from AB17: {e}")

    for name, content in all_contents:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]; cat = identify_category(ws)
            if cat:
                proc = PROCESSORS.get(cat)
                res = proc.process_sheet(ws, name, master_equip, equip_order=equip_order, ab17_cable_data=ab17_cable_data)
                if res:
                    # Clean and validate bidder name
                    b_raw = res.get("BidderRaw", "")
                    issues_list = res.get("Issues") if isinstance(res.get("Issues"), list) else [res.get("Issues")]
                    issues_list = [iss for iss in issues_list if str(iss).strip()]
                    missing_bidder_issue = any(
                        isinstance(iss, str) and "Forgot to fill bidder in file" in iss
                        for iss in issues_list
                    )

                    if cat == "AB17" and not is_valid_bidder(b_raw):
                        res["BidderRaw"] = ""
                        if not missing_bidder_issue:
                            issues_list.append(f"Forgot to fill bidder in file {name}")
                        res["Issues"] = issues_list
                        res["Status"] = "INVALID"
                    elif not missing_bidder_issue and not is_valid_bidder(b_raw):
                        fallback = os.path.splitext(name)[0]
                        fallback = re.sub(r'(_)?(AB16|AB17|AB18)(_)?', '', fallback, flags=re.I)
                        fallback = re.sub(r'\b(PROPOSAL|SCHEDULE|DATA|TEMPLATE)\b', '', fallback, flags=re.I)
                        fallback = re.sub(r'[\-_]+', ' ', fallback)
                        fallback = " ".join(fallback.split()).strip()
                        res["BidderRaw"] = fallback or "Unknown Bidder"

                    item_no_raw = str(res.get("ItemNo") or "").strip()
                    if item_no_raw:
                        it_id = item_no_raw.split()[0].upper()
                        if re.search(r"\d+AB16", it_id): res["Category"] = "AB16_CLEAT" if ("CLEAT" in it_id or cat == "AB16_CLEAT") else "AB16_TERM"
                        elif re.search(r"\d+AB17", it_id): res["Category"] = "AB17"
                        elif re.search(r"\d+AB18", it_id): res["Category"] = "AB18"
                        else: res["Category"] = cat
                        raw_results.append(res)
                    else:
                        res["Category"] = cat
                        raw_results.append(res)

    # 3. Grouping
    bidders_map = {}
    master_bidder_names = [] # To keep track of bidders across ALL categories
    skipped_results = []
    groupable_results = []
    
    print(f"[DEBUG] Starting grouping for {len(raw_results)} results")
    
    for r in raw_results:
        b_raw = r["BidderRaw"]
        issues_list = r.get("Issues") if isinstance(r.get("Issues"), list) else [r.get("Issues")]
        missing_bidder_issue = any(
            isinstance(iss, str) and "Forgot to fill bidder in file" in iss
            for iss in issues_list
        )
        if missing_bidder_issue or not str(b_raw or "").strip():
            skipped_issues = sorted(list({str(iss).strip() for iss in issues_list if str(iss).strip()}))
            if not skipped_issues:
                skipped_issues = [f"Forgot to fill bidder in file {r.get('SourceFile', 'Unknown file')}"]
            skipped_results.append({
                "Bidder": "",
                "Category": r.get("Category", "UNKNOWN"),
                "Status": "MISSING_BIDDER",
                "Issues": skipped_issues,
                "Schedule": r.get("Schedule", ""),
                "Equipment": {},
                "equip_order": [],
                "all_equip_order": equip_order,
                "ItemNoSummary": "",
            })
            print(f"[DEBUG] Skipping validation for missing bidder in category '{r.get('Category', 'UNKNOWN')}'")
            continue
        groupable_results.append(r)

    # Pass 1: Normalize all bidder names globally
    for r in groupable_results:
        b_raw = r["BidderRaw"]
        b_clean = robust_clean(b_raw).upper()
        b_norm = BIDDER_ALIASES.get(b_clean, b_clean)
        
        # Fuzzy match against global list to unify names across categories
        found = None
        if master_bidder_names:
            best_match, score = fuzz_process.extractOne(b_norm, master_bidder_names)
            if score >= BIDDER_MATCH_THRESHOLD:
                found = best_match
        
        target_name = found or b_norm
        if target_name not in master_bidder_names:
            master_bidder_names.append(target_name)
        
        r["NormalizedBidder"] = target_name
        print(f"[DEBUG] Global Normalization: '{b_raw}' -> '{target_name}'")

    # Pass 2: Group by normalized name and category
    for r in groupable_results:
        target = r["NormalizedBidder"]
        cat = r.get("Category", "UNKNOWN")
        key = f"{target}|{cat}"
        
        if key not in bidders_map:
            print(f"[DEBUG] Creating new entry for key: {key}")
            prefix = cat.split("_")[0]
            if prefix == "AB16":
                if "CLEAT" in cat:
                    cat_keys = [eq for eq in equip_order if prefix in eq.split()[0].upper() and "CLEAT" in eq.upper()]
                else:
                    cat_keys = [eq for eq in equip_order if prefix in eq.split()[0].upper() and "CLEAT" not in eq.upper()]
            else:
                cat_keys = [eq for eq in equip_order if prefix in eq.split()[0].upper()]
            
            bidders_map[key] = {
                "Bidder": target, "Category": cat, "Status": "VALID", "Issues": set(), "Schedule": r.get("Schedule", ""),
                "Equipment": {eq: [] for eq in cat_keys}, "equip_order": cat_keys, "all_equip_order": equip_order, "ItemNos": set()
            }
        
        ent = bidders_map[key]
        if r.get("Status") != "VALID": ent["Status"] = r.get("Status")
        for iss in (r.get("Issues") if isinstance(r.get("Issues"), list) else [r.get("Issues")]):
            if str(iss).strip(): ent["Issues"].add(str(iss).strip())
        for it in str(r.get("ItemNo") or "").split(", "):
            it_c = it.strip().upper()
            if it_c and not any(t in it_c for t in UNWANTED_TERMS): ent["ItemNos"].add(it_c)
        for eq, rows in r.get("Equipment", {}).items():
            if eq in ent["Equipment"]:
                for row in rows:
                    if row not in ent["Equipment"][eq]: ent["Equipment"][eq].append(row)

    final = list(skipped_results)
    for b in bidders_map.values():
        b["ItemNoSummary"] = ", ".join(sorted(list(b["ItemNos"])))
        missing = [eq for eq in b["equip_order"] if eq.split()[0] not in b["ItemNos"]]
        if missing:
            b["Issues"].add(f"Missing items: {', '.join(missing)}")
            if b["Status"] == "VALID": b["Status"] = "INCOMPLETE"
        b["Issues"] = sorted(list(b["Issues"])); out = b.copy(); del out["ItemNos"]; final.append(out)
    return {"data": clean_surrogates(final)}

@app.post("/generate-report")
async def generate_report(data: List[Dict], category: str = "ALL"):
    from .services.reporter import Reporter
    return Reporter().generate(data, category)

if __name__ == "__main__":
    import uvicorn; uvicorn.run(app, host="0.0.0.0", port=8888)

