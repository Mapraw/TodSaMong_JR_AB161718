import base64
import io
import re
import sys
import traceback
from pathlib import Path
from typing import List

import openpyxl
import pandas as pd
import uvicorn
from fastapi import FastAPI, File, HTTPException, UploadFile
from openpyxl.styles import Alignment, Font, PatternFill

CURRENT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = CURRENT_DIR.parent
PROJECT_DIR = BACKEND_DIR.parent
for candidate in (CURRENT_DIR, BACKEND_DIR, PROJECT_DIR):
    candidate_str = str(candidate)
    if candidate_str not in sys.path:
        sys.path.append(candidate_str)

try:
    from backend.ab17.pd_analyzer import PDAnalyzer
    from backend.shared.proposal_helpers import (
        extract_detailed_manufacturer_from_pd,
        extract_item_id_from_pd,
        extract_metadata_from_price_schedule,
        normalize_id,
    )
    from backend.shared.standard_paths import category_standards_dir
except ImportError:
    from ..ab17.pd_analyzer import PDAnalyzer
    from ..shared.proposal_helpers import (
        extract_detailed_manufacturer_from_pd,
        extract_item_id_from_pd,
        extract_metadata_from_price_schedule,
        normalize_id,
    )
    from ..shared.standard_paths import category_standards_dir

app = FastAPI(title="Thriller API v2.1 - UPDATED")


@app.get("/")
async def root():
    return {"message": "Thriller API is running on port 8890"}


print("\n" + "=" * 50)
print("THRILLER BACKEND v2.1 STARTED - GEMINI FIXED")
print("=" * 50 + "\n")
print("\n" + "!" * 60)
print("!!! THRILLER BACKEND UPDATED BY GEMINI - VERSION 2.5 !!!")
print("!!! PLEASE RESTART SERVER IF YOU DON'T SEE THIS !!!")
print("!" * 60 + "\n")

STANDARDS_DIR = category_standards_dir("ab17", fallback_to_legacy=True)
pd_analyzer = PDAnalyzer(STANDARDS_DIR)


@app.post("/analyze")
async def analyze_documents(
    proposal_excels: List[UploadFile] = File(...),
    price_schedule: UploadFile = File(None),
):
    try:
        price_map = {}
        if price_schedule:
            price_map = extract_metadata_from_price_schedule(await price_schedule.read())

        section_names = {
            "e": "Area",
            "f": "Diameter",
            "g": "Cond. Screen",
            "h": "Insulation",
            "i": "Ins. Screen",
            "j": "OD Insulation",
            "k": "Metal Screen",
            "l": "Cushion",
            "m": "Oversheath",
            "n": "Overall Diameter",
        }

        final_table_data = []
        for pd_file in proposal_excels:
            excel_content = await pd_file.read()
            df_pd = pd.read_excel(io.BytesIO(excel_content), header=None)
            raw_item_id = extract_item_id_from_pd(df_pd)
            clean_item_key = normalize_id(raw_item_id)

            metadata = price_map.get(clean_item_key)
            if metadata:
                sch_no = metadata["schedule"]
                item_no = metadata["item_no"]
                item_desc = metadata["description"]
            else:
                schedule_match = re.search(r"\d+AB\d+", raw_item_id, re.I) if raw_item_id else None
                sch_no = schedule_match.group(0).upper() if schedule_match else "Unknown"
                item_no = raw_item_id or "Unknown"
                item_desc = "Description Not Found"

            mfr_data = extract_detailed_manufacturer_from_pd(df_pd)
            mfr_display = mfr_data.get("manufacturer", "")
            if mfr_data.get("country"):
                mfr_display += f" - {mfr_data['country']}"
            if mfr_data.get("type_model"):
                mfr_display += f" | {mfr_data['type_model']}"

            analysis_results = pd_analyzer.analyze_sections(df_pd)

            error_sections = []
            detailed_comments = []
            for sec_key in "efghijklmn":
                if sec_key not in analysis_results:
                    continue
                result = analysis_results[sec_key]
                if result.get("status") == "ผิด":
                    section_name = section_names.get(sec_key, sec_key)
                    error_sections.append(f"❌ {sec_key}. {section_name}")
                    detailed_comments.append(f"• {sec_key}: {result.get('comment', '')}")

            if not error_sections:
                final_status = "✅ ผ่านทุกหัวข้อ (e-n)"
                final_comment = "ตรวจสอบแล้วถูกต้องครบถ้วน"
            else:
                final_status = "\n".join(error_sections)
                final_comment = "พบจุดที่ต้องแก้ไข:\n" + "\n".join(detailed_comments)

            final_table_data.append(
                {
                    "Schedule": sch_no,
                    "Item No.": item_no,
                    "Description": item_desc,
                    "Manufacturer": mfr_display,
                    "ผลการพิจารณา": final_status,
                    "ความคิดเห็นเพิ่มเติม": final_comment,
                }
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        headers = [
            "Schedule",
            "Item No.",
            "Description",
            "Manufacturer",
            "ผลการพิจารณา",
            "ความคิดเห็นเพิ่มเติม",
        ]
        for column_index, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for row_index, row in enumerate(final_table_data, start=2):
            worksheet.cell(row=row_index, column=1, value=row["Schedule"]).alignment = Alignment(vertical="top")
            worksheet.cell(row=row_index, column=2, value=row["Item No."]).alignment = Alignment(vertical="top")
            worksheet.cell(row=row_index, column=3, value=row["Description"]).alignment = Alignment(
                wrapText=True,
                vertical="top",
            )

            manufacturer_cell = worksheet.cell(row=row_index, column=4, value=row["Manufacturer"])
            manufacturer_cell.alignment = Alignment(wrapText=True, vertical="top")

            status_cell = worksheet.cell(row=row_index, column=5, value=row["ผลการพิจารณา"])
            fill_color = "C6EFCE" if "ผ่าน" in row["ผลการพิจารณา"] else "FFC7CE"
            status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            status_cell.alignment = Alignment(wrapText=True, vertical="top")

            worksheet.cell(row=row_index, column=6, value=row["ความคิดเห็นเพิ่มเติม"]).alignment = Alignment(
                wrapText=True,
                vertical="top",
            )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        excel_base64 = base64.b64encode(output.read()).decode("utf-8")

        return {
            "specialized_results": final_table_data,
            "excel_base64": excel_base64,
        }
    except Exception as exc:
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(exc))


if __name__ == "__main__":
    uvicorn.run("backend.thriller.thriller_main:app", host="127.0.0.1", port=8890, reload=True)
