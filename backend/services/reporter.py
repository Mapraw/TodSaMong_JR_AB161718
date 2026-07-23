import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from typing import List, Dict
from fastapi.responses import StreamingResponse

class Reporter:
    def generate(self, data: List[Dict], category: str = "ALL"):
        try:
            if category == "AB17": return self._generate_ab17_report(data)
            elif category == "AB18": return self._generate_ab18_report(data)
            elif category == "AB16": return self._generate_ab16_report(data)
            else: return self._generate_overall_consolidated_report(data)
        except Exception as e:
            print(f"[REPORTER ERROR] {e}")
            import traceback
            traceback.print_exc()
            return {"error": str(e)}

    def _generate_overall_consolidated_report(self, data: List[Dict]):
        output = io.BytesIO()
        workbook = pd.ExcelWriter(output, engine='xlsxwriter')
        wb_obj = workbook.book
        
        # Formats
        head_fmt = wb_obj.add_format({'bold': True, 'bg_color': '#FFD700', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
        wrap_fmt = wb_obj.add_format({'text_wrap': True, 'valign': 'top', 'border': 1, 'font_name': 'Arial'})
        pass_fmt = wb_obj.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'border': 1, 'font_name': 'Arial', 'align': 'center', 'valign': 'top', 'text_wrap': True})
        fail_fmt = wb_obj.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'border': 1, 'font_name': 'Arial', 'align': 'center', 'valign': 'top', 'text_wrap': True})
        warn_fmt = wb_obj.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C6500', 'border': 1, 'font_name': 'Arial', 'align': 'center', 'valign': 'top', 'text_wrap': True})

        # Extended unwanted terms to be very strict
        unwanted_terms = [
            "TOTAL PRICE", "SUMMARY", "TRANSPORTATION", "GRAND TOTAL", 
            "TOTAL PRICE FOR SCHEDULE", "CONSTRUCTION AND INSTALLATION",
            "PRICE OF SCHEDULE", "PRICE FOR SCHEDULE"
        ]

        # --- 1. Collected Unique Data ---
        all_equip_keys = []
        seen_equip = set()
        unique_bidders = sorted(list(set(b.get("Bidder") for b in data)))
        
        # Consistent order for equipment extraction
        for b in sorted(data, key=lambda x: x.get("Category", "")):
            # Check all possible sources of item keys
            sources = (b.get("all_equip_order") or []) + (b.get("equip_order") or []) + list(b.get("Equipment", {}).keys())
            for eq_key in sources:
                if eq_key not in seen_equip:
                    eq_upper = eq_key.upper()
                    if not any(term in eq_upper for term in unwanted_terms):
                        all_equip_keys.append(eq_key); seen_equip.add(eq_key)
        
        # Sort equipment keys by Item ID naturally (extracting the ID part correctly)
        def sort_key(s):
            match = re.search(r"\d+AB\d+-\d+", s, re.I)
            return match.group(0).upper() if match else s
        all_equip_keys.sort(key=sort_key)

        # --- 2. Detailed Summary Sheet (Comprehensive Matrix) ---
        # Matrix: [Item ID, Description] + [Bidder 1 Manuf, Bidder 1 Result, Bidder 2 Manuf, Bidder 2 Result...]
        summary_rows = []
        for eq_key in all_equip_keys:
            eq_id = eq_key.split()[0]
            eq_desc = " ".join(eq_key.split()[1:])
            row = [eq_id, eq_desc]
            for bidder_name in unique_bidders:
                bidder_entries = [b for b in data if b.get("Bidder") == bidder_name]
                found_in_any = False
                mfr_text = ""
                res_text = "N/A"
                
                for entry in bidder_entries:
                    if eq_key in entry.get("Equipment", {}):
                        found_in_any = True
                        mfrs = entry["Equipment"][eq_key]
                        if mfrs:
                            mfr_list = []
                            res_list = []
                            for m in mfrs:
                                m_name = m.get("manufacturer", "Unknown")
                                mfr_list.append(m_name)
                                res_list.append(m.get("result", "N/A"))
                            mfr_text = "\n".join([f"- {m}" for m in mfr_list])
                            all_pass = all(any(ok in str(r).upper() for ok in ["PASS", "VALID", "ผ่าน"]) and "ไม่ผ่าน" not in str(r) for r in res_list)
                            res_text = "Pass" if all_pass else "Fail"
                        break
                
                if not found_in_any:
                    row.extend(["", "NOT OFFERED"])
                else:
                    row.extend([mfr_text, res_text])
            summary_rows.append(row)

        header = ["Item ID", "Description"]
        for b in unique_bidders: 
            header.extend([f"{b}\nManufacturer", f"{b}\nResult"])
        
        df_sum = pd.DataFrame(summary_rows, columns=header)
        df_sum.to_excel(workbook, sheet_name="Comprehensive Summary", index=False)
        ws_sum = workbook.sheets["Comprehensive Summary"]
        for col_num, val in enumerate(header): ws_sum.write(0, col_num, val, head_fmt)
        ws_sum.set_column('A:A', 15); ws_sum.set_column('B:B', 40, wrap_fmt)
        for i in range(len(unique_bidders)):
            ws_sum.set_column(2 + i*2, 2 + i*2, 25, wrap_fmt) # Manuf
            ws_sum.set_column(3 + i*2, 3 + i*2, 12, wrap_fmt) # Result
            
        for r_idx, row in enumerate(summary_rows):
            for i in range(len(unique_bidders)):
                mfr_val = row[2 + i*2]
                res_val = row[3 + i*2]
                ws_sum.write(r_idx + 1, 2 + i*2, mfr_val, wrap_fmt)
                fmt = pass_fmt if res_val == "Pass" else fail_fmt if res_val == "Fail" else warn_fmt
                ws_sum.write(r_idx + 1, 3 + i*2, res_val, fmt)

        # --- 3. Per-Bidder Sheets ---
        for bidder_name in unique_bidders:
            bidder_entries = [b for b in data if b.get("Bidder") == bidder_name]
            ws_name = re.sub(r'[\[\]\*\?\:\\/]', '', bidder_name)[:31]
            
            # Consolidated list of unique schedules for this bidder
            all_schedules = set()
            for b in bidder_entries:
                sch = str(b.get("Schedule", "")).strip()
                if sch: all_schedules.add(sch)
            
            rows = [
                ["GENERAL INFORMATION", ""], ["Bidder Name", bidder_name],
                ["Schedules", ", ".join(sorted(list(all_schedules)))],
                ["Overall Status", "PASS" if not any(b.get("Status") in ["INVALID", "INCOMPLETE"] for b in bidder_entries) else "FAIL"],
                ["Issues", "\n".join([f"- {iss}" for iss in sorted(list(set(issue for b in bidder_entries for issue in b.get("Issues", []))))])] ,
                ["", ""], ["EQUIPMENT SPECIFICATIONS", "MANUFACTURER", "COUNTRY", "STANDARDS", "DETAILS", "RESULT", "COMMENT"]
            ]
            header_row_count = len(rows)
            for eq_key in all_equip_keys:
                found_in_any = False
                for entry in bidder_entries:
                    if eq_key in entry.get("Equipment", {}):
                        found_in_any = True
                        mfrs = entry["Equipment"][eq_key]
                        if not mfrs:
                            rows.append([eq_key, "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"])
                        else:
                            for m in mfrs:
                                details = "\n".join([f"- {k}: {m[k]}" for k in ["type_model", "od_range", "material_class", "formation", "peak_current"] if m.get(k)])
                                rows.append([eq_key, m.get("manufacturer"), m.get("country"), m.get("standard"), details, m.get("result"), m.get("comment")])
                        break
                if not found_in_any:
                    rows.append([eq_key, "", "", "", "", "NOT OFFERED", "Item not found in proposal data"])

            pd.DataFrame(rows).to_excel(workbook, sheet_name=ws_name, index=False, header=False)
            ws = workbook.sheets[ws_name]
            ws.set_column('A:A', 30, wrap_fmt); ws.set_column('B:D', 20, wrap_fmt); ws.set_column('E:G', 30, wrap_fmt)
            ws.write('A1', "GENERAL INFORMATION", head_fmt); ws.write('A7', "EQUIPMENT SPECIFICATIONS", head_fmt)
            for col in range(1, 7): ws.write(6, col, rows[6][col], head_fmt)
            for r_idx in range(header_row_count, len(rows)):
                res_val = str(rows[r_idx][5])
                fmt = pass_fmt if (any(ok in res_val.upper() for ok in ["PASS", "VALID", "ผ่าน"]) and "ไม่ผ่าน" not in res_val) else fail_fmt if (any(no in res_val.upper() for no in ["FAIL", "INVALID", "ผิด"]) or "ไม่ผ่าน" in res_val) else warn_fmt
                ws.write(r_idx, 5, res_val, fmt)

        workbook.close()
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=AB16-AB18_{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"})

    def _generate_ab17_report(self, data: List[Dict]):
        # Specific AB17 report (original style)
        new_wb = openpyxl.Workbook(); ws = new_wb.active; ws.title = "AB17 Analysis"
        headers = ["Schedule", "Item No.", "Description", "Manufacturer", "Country", "Type/Model", "ผลการพิจารณา", "ความคิดเห็นเพิ่มเติม"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row_idx = 2
        for bidder in sorted(data, key=lambda x: x.get("Bidder", "")):
            for eq_key, mfrs in bidder.get("Equipment", {}).items():
                if "AB17" not in str(eq_key): continue
                for m_data in mfrs:
                    ws.cell(row=row_idx, column=1, value=bidder.get("Schedule")).alignment = Alignment(vertical="top")
                    ws.cell(row=row_idx, column=2, value=bidder.get("ItemNoSummary")).alignment = Alignment(vertical="top")
                    ws.cell(row=row_idx, column=3, value=eq_key).alignment = Alignment(wrapText=True, vertical="top")
                    
                    cell_manuf = ws.cell(row=row_idx, column=4, value=m_data.get("manufacturer", ""))
                    cell_manuf.alignment = Alignment(wrapText=True, vertical="top")
                    
                    cell_country = ws.cell(row=row_idx, column=5, value=m_data.get("country", ""))
                    cell_country.alignment = Alignment(wrapText=True, vertical="top")
                    
                    cell_model = ws.cell(row=row_idx, column=6, value=m_data.get("type_model", ""))
                    cell_model.alignment = Alignment(wrapText=True, vertical="top")
                    
                    res = str(m_data.get("result", "")); cell_status = ws.cell(row=row_idx, column=7, value=res)
                    color = "C6EFCE" if (any(x in res.upper() for x in ["PASS", "VALID", "ผ่าน"]) and "ไม่ผ่าน" not in res) else "FFC7CE"
                    cell_status.fill = PatternFill(start_color=color, end_color=color, fill_type="solid"); cell_status.alignment = Alignment(wrapText=True, vertical="top")
                    
                    ws.cell(row=row_idx, column=8, value=m_data.get("comment", "")).alignment = Alignment(wrapText=True, vertical="top")
                    row_idx += 1
        output = io.BytesIO(); new_wb.save(output); output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=AB17_Report.xlsx"})

    def _generate_ab18_report(self, data: List[Dict]):
        output = io.BytesIO(); ab18_data = [b for b in data if b["Category"] == "AB18"]
        if not ab18_data: return {"error": "No AB18 data"}
        workbook = pd.ExcelWriter(output, engine='xlsxwriter'); self._write_ab18_sheets(workbook, ab18_data); workbook.close(); output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=AB18_Report.xlsx"})

    def _write_ab18_sheets(self, workbook, data, prefix=""):
        wb_obj = workbook.book
        head_fmt = wb_obj.add_format({'bold': True, 'bg_color': '#FFD700', 'border': 1})
        pass_fmt = wb_obj.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'border': 1, 'font_name': 'Arial', 'text_wrap': True, 'valign': 'top'})
        fail_fmt = wb_obj.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'border': 1, 'font_name': 'Arial', 'text_wrap': True, 'valign': 'top'})
        sorted_data = sorted(data, key=lambda x: x.get("Bidder", ""))
        equip_order = sorted_data[0].get("equip_order", []) if sorted_data else []
        bidders = [b.get("Bidder") for b in sorted_data]
        header = ["Item ID", "Description"] + bidders
        summary_rows = []
        for eq_key in equip_order:
            if "AB18" not in eq_key: continue
            row = [eq_key.split()[0], " ".join(eq_key.split()[1:])]
            for b in sorted_data:
                res_list = b.get("Equipment", {}).get(eq_key, [])
                all_pass = all(any(ok in str(r.get("result", "")).upper() for ok in ["PASS", "VALID", "ผ่าน"]) and "ไม่ผ่าน" not in str(r.get("result", "")) for r in res_list) if res_list else False
                row.append("Pass" if all_pass else "Fail" if res_list else "N/A")
            summary_rows.append(row)
        ws_name = f"{prefix}AB18 Summary"[:31]
        pd.DataFrame(summary_rows, columns=header).to_excel(workbook, sheet_name=ws_name, index=False)
        ws = workbook.sheets[ws_name]
        for col_num, val in enumerate(header): ws.write(0, col_num, val, head_fmt)
        for r_idx, row in enumerate(summary_rows):
            for c_idx, val in enumerate(row[2:], 2): ws.write(r_idx + 1, c_idx, val, pass_fmt if val == "Pass" else fail_fmt)

    def _generate_ab16_report(self, data: List[Dict]):
        output = io.BytesIO(); ab16_data = [b for b in data if b["Category"].startswith("AB16")]
        if not ab16_data: return {"error": "No AB16 data"}
        workbook = pd.ExcelWriter(output, engine='xlsxwriter'); self._write_ab16_sheets(workbook, ab16_data); workbook.close(); output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=AB16_Report.xlsx"})

    def _write_ab16_sheets(self, workbook, data, prefix=""):
        wb_obj = workbook.book
        head_fmt = wb_obj.add_format({'bold': True, 'bg_color': '#FFD700', 'border': 1})
        sorted_data = sorted(data, key=lambda x: x.get("Bidder", ""))
        equip_keys = set()
        for b in sorted_data:
            for k in b.get("Equipment", {}).keys():
                if "AB16" in str(k): equip_keys.add(k)
        equip_order = sorted(list(equip_keys))
        header = ["Item ID", "Description"] + [b.get("Bidder") for b in sorted_data]
        summary_rows = []
        for eq_key in equip_order:
            row = [eq_key.split()[0], " ".join(eq_key.split()[1:])]
            for b in sorted_data:
                res_list = b.get("Equipment", {}).get(eq_key, [])
                all_pass = all(any(ok in str(r.get("result", "")).upper() for ok in ["PASS", "VALID", "ผ่าน"]) and "ไม่ผ่าน" not in str(r.get("result", "")) for r in res_list) if res_list else False
                row.append("Pass" if all_pass else "Fail" if res_list else "N/A")
            summary_rows.append(row)
        ws_name = f"{prefix}AB16 Summary"[:31]
        pd.DataFrame(summary_rows, columns=header).to_excel(workbook, sheet_name=ws_name, index=False)
